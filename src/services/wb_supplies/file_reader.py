from __future__ import annotations

import csv
import importlib
import io
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable, Sequence

from openpyxl import load_workbook

from src.clients.google_drive_client import (
    CSV_MIME_TYPES,
    GOOGLE_SHEETS_MIME_TYPE,
    GOOGLE_SHORTCUT_MIME_TYPE,
    MS_EXCEL_MIME_TYPES,
    GoogleDriveClient,
    get_effective_mime_type,
    get_unsupported_supply_reason,
)
from src.clients.google_sheets_client import GoogleSheetsClient

CSV_ENCODINGS = ("utf-8-sig", "utf-8", "cp1251")
XLSX_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
XLS_EXTENSIONS = {".xls"}
CSV_EXTENSIONS = {".csv"}


@dataclass(slots=True)
class ParsedWorksheet:
    sheet_name: str
    rows: list[list[Any]]
    raw_rows_count: int
    detected_header_candidates: list[dict[str, Any]] = field(default_factory=list)
    preview_rows: list[list[Any]] = field(default_factory=list)


@dataclass(slots=True)
class ParsedWorkbook:
    source_file_id: str
    source_file_name: str
    source_mime_type: str | None
    effective_file_id: str
    effective_file_name: str
    effective_mime_type: str | None
    worksheets: list[ParsedWorksheet]
    warnings: list[str] = field(default_factory=list)
    shortcut_source: dict[str, Any] | None = None


HeaderDetector = Callable[[Sequence[Sequence[Any]]], tuple[int | None, dict[str, int]]]


def _sheet_range(sheet_name: str) -> str:
    escaped = sheet_name.replace("'", "''")
    return f"'{escaped}'!A:ZZ"


def _file_extension(file_item: dict[str, Any]) -> str:
    return Path(str(file_item.get("name") or "")).suffix.lower()


def _is_csv_file(file_item: dict[str, Any]) -> bool:
    mime_type = file_item.get("mimeType")
    extension = _file_extension(file_item)
    return mime_type in CSV_MIME_TYPES or extension in CSV_EXTENSIONS


def _is_xls_file(file_item: dict[str, Any]) -> bool:
    return _file_extension(file_item) in XLS_EXTENSIONS


def _is_xlsx_file(file_item: dict[str, Any]) -> bool:
    mime_type = file_item.get("mimeType")
    extension = _file_extension(file_item)
    if extension in XLSX_EXTENSIONS:
        return True
    return mime_type in MS_EXCEL_MIME_TYPES and not _is_xls_file(file_item) and not _is_csv_file(file_item)


def _detect_header_candidates(
    rows: Sequence[Sequence[Any]],
    header_detector: HeaderDetector | None,
) -> list[dict[str, Any]]:
    if not header_detector:
        return []
    header_row_index, mapping = header_detector(rows)
    if header_row_index is None:
        return []
    header_row = list(rows[header_row_index]) if header_row_index < len(rows) else []
    return [{
        "row_number": header_row_index + 1,
        "mapping": dict(mapping),
        "header_preview": header_row[:12],
    }]


def _read_google_sheet(
    file_item: dict[str, Any],
    sheets_client: GoogleSheetsClient,
    header_detector: HeaderDetector | None,
) -> list[ParsedWorksheet]:
    spreadsheet_id = str(file_item.get("id") or "")
    worksheet_titles = sheets_client.get_worksheet_titles(spreadsheet_id) or []
    if not worksheet_titles:
        raise RuntimeError("No worksheet titles returned.")

    worksheets: list[ParsedWorksheet] = []
    for sheet_name in worksheet_titles:
        rows = [list(row) for row in (sheets_client.read_range(spreadsheet_id, _sheet_range(sheet_name)) or [])]
        worksheets.append(
            ParsedWorksheet(
                sheet_name=sheet_name,
                rows=rows,
                raw_rows_count=len(rows),
                detected_header_candidates=_detect_header_candidates(rows, header_detector),
                preview_rows=rows[:5],
            )
        )
    return worksheets


def _read_xlsx_bytes(file_bytes: bytes, header_detector: HeaderDetector | None) -> list[ParsedWorksheet]:
    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    worksheets: list[ParsedWorksheet] = []
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
        if not rows:
            continue
        worksheets.append(
            ParsedWorksheet(
                sheet_name=sheet_name,
                rows=rows,
                raw_rows_count=len(rows),
                detected_header_candidates=_detect_header_candidates(rows, header_detector),
                preview_rows=rows[:5],
            )
        )
    return worksheets


def _read_xls_bytes(file_bytes: bytes, header_detector: HeaderDetector | None) -> list[ParsedWorksheet]:
    try:
        xlrd = importlib.import_module("xlrd")
    except ImportError as exc:
        raise RuntimeError("Файл .xls найден, но парсер xlrd не установлен / не поддержан.") from exc

    workbook = xlrd.open_workbook(file_contents=file_bytes)
    worksheets: list[ParsedWorksheet] = []
    for sheet in workbook.sheets():
        rows = [sheet.row_values(row_index) for row_index in range(sheet.nrows)]
        if not rows:
            continue
        worksheets.append(
            ParsedWorksheet(
                sheet_name=sheet.name,
                rows=rows,
                raw_rows_count=len(rows),
                detected_header_candidates=_detect_header_candidates(rows, header_detector),
                preview_rows=rows[:5],
            )
        )
    return worksheets


def _read_csv_bytes(file_bytes: bytes, file_name: str) -> tuple[list[ParsedWorksheet], list[str]]:
    warnings: list[str] = []
    last_error: str | None = None
    for encoding in CSV_ENCODINGS:
        try:
            text = file_bytes.decode(encoding)
            rows = [list(row) for row in csv.reader(io.StringIO(text))]
            return [
                ParsedWorksheet(
                    sheet_name=Path(file_name).stem or "CSV",
                    rows=rows,
                    raw_rows_count=len(rows),
                    detected_header_candidates=[],
                    preview_rows=rows[:5],
                )
            ], warnings
        except UnicodeDecodeError as exc:
            last_error = str(exc)
            warnings.append(f"{file_name}: csv decode failed for {encoding}")
    raise RuntimeError(f"Не удалось декодировать CSV-файл: {last_error or 'unknown error'}")


def read_supply_file(
    file_item: dict[str, Any],
    *,
    drive_client: GoogleDriveClient,
    sheets_client: GoogleSheetsClient,
    header_detector: HeaderDetector | None = None,
) -> ParsedWorkbook:
    unsupported_reason = get_unsupported_supply_reason(file_item)
    if unsupported_reason:
        return ParsedWorkbook(
            source_file_id=str(file_item.get("id") or ""),
            source_file_name=str(file_item.get("name") or ""),
            source_mime_type=file_item.get("mimeType"),
            effective_file_id=str(file_item.get("id") or ""),
            effective_file_name=str(file_item.get("name") or ""),
            effective_mime_type=get_effective_mime_type(file_item),
            worksheets=[],
            warnings=[unsupported_reason],
        )

    effective_file = drive_client.resolve_shortcut_target(file_item)
    effective_mime_type = effective_file.get("mimeType")
    effective_file_id = str(effective_file.get("id") or "")
    effective_file_name = str(effective_file.get("name") or effective_file_id)
    warnings: list[str] = []

    if effective_mime_type == GOOGLE_SHEETS_MIME_TYPE:
        worksheets = _read_google_sheet(effective_file, sheets_client, header_detector)
    else:
        file_bytes = drive_client.download_file_bytes(effective_file_id)
        if _is_csv_file(effective_file):
            worksheets, csv_warnings = _read_csv_bytes(file_bytes, effective_file_name)
            warnings.extend(csv_warnings)
        elif _is_xls_file(effective_file):
            worksheets = _read_xls_bytes(file_bytes, header_detector)
        elif _is_xlsx_file(effective_file):
            worksheets = _read_xlsx_bytes(file_bytes, header_detector)
        else:
            raise RuntimeError(
                f"Unsupported effective mimeType for supply reader: {effective_mime_type or 'unknown'}"
            )

    for worksheet in worksheets:
        if not worksheet.detected_header_candidates:
            worksheet.detected_header_candidates = _detect_header_candidates(worksheet.rows, header_detector)

    return ParsedWorkbook(
        source_file_id=str(file_item.get("id") or ""),
        source_file_name=str(file_item.get("name") or effective_file_name),
        source_mime_type=file_item.get("mimeType"),
        effective_file_id=effective_file_id,
        effective_file_name=effective_file_name,
        effective_mime_type=effective_mime_type,
        worksheets=worksheets,
        warnings=warnings,
        shortcut_source=effective_file.get("shortcutSource"),
    )