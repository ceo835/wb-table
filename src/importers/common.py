from __future__ import annotations

import re
from io import BytesIO
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.cell import Cell, ReadOnlyCell
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


DATE_PATTERNS = (
    re.compile(r"(?P<iso>\d{4}-\d{2}-\d{2})"),
    re.compile(r"(?P<ru>\d{2}\.\d{2}\.\d{4})"),
)


@dataclass(frozen=True)
class SheetSelection:
    sheet_name: str | None
    header_row_index: int | None
    headers: tuple[str, ...]
    missing_required_columns: tuple[str, ...]
    overlap_count: int = 0


@dataclass(frozen=True)
class ParsedImportResult:
    file_path: Path
    detected_date: date | None
    sheet_name: str | None
    rows_read: int
    rows_normalized: tuple[dict[str, Any], ...]
    missing_required_columns: tuple[str, ...]
    skipped_rows_count: int = 0
    skipped_rows_preview: tuple[dict[str, Any], ...] = ()


def load_workbook_readonly(file_path: str | Path) -> Workbook:
    file_bytes = Path(file_path).read_bytes()
    # Some WB exports have broken sheet dimensions in read_only mode and expose only the first column.
    return load_workbook(filename=BytesIO(file_bytes), data_only=True, read_only=False)


def close_workbook(workbook: Workbook | None) -> None:
    if workbook is None:
        return
    close_method = getattr(workbook, "close", None)
    if callable(close_method):
        close_method()


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").replace("\r", " ").replace("\xa0", " ")
    return re.sub(r"\s+", " ", text).strip()


def normalize_text(value: Any) -> str | None:
    text = normalize_header(value)
    return text or None


def parse_date_text(value: str) -> date | None:
    for pattern in DATE_PATTERNS:
        match = pattern.search(value)
        if not match:
            continue
        raw = match.group("iso") if "iso" in match.groupdict() and match.group("iso") else match.group("ru")
        if not raw:
            continue
        if "." in raw:
            return datetime.strptime(raw, "%d.%m.%Y").date()
        return date.fromisoformat(raw)
    return None


def detect_date_from_filename(file_path: str | Path) -> date | None:
    return parse_date_text(Path(file_path).name)


def detect_date_from_info_sheet(workbook: Workbook) -> date | None:
    for sheet in workbook.worksheets:
        title = normalize_header(sheet.title).lower()
        if "общая" not in title or "информация" not in title:
            continue
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 30), values_only=True):
            for value in row:
                if value is None:
                    continue
                parsed = parse_date_text(str(value))
                if parsed is not None:
                    return parsed
    return None


def resolve_import_date(file_path: str | Path, explicit_date: date | None = None, workbook: Workbook | None = None) -> date | None:
    if explicit_date is not None:
        return explicit_date
    if workbook is not None:
        workbook_date = detect_date_from_info_sheet(workbook)
        if workbook_date is not None:
            return workbook_date
    return detect_date_from_filename(file_path)


def _row_values(sheet: Worksheet, row_index: int) -> tuple[Any, ...]:
    return tuple(cell.value for cell in next(sheet.iter_rows(min_row=row_index, max_row=row_index)))


def select_sheet_by_required_columns(
    workbook: Workbook,
    required_columns: Iterable[str],
    preferred_sheet_name: str | None = None,
    max_header_row: int = 6,
) -> SheetSelection:
    required = tuple(required_columns)
    required_set = set(required)

    def evaluate_sheet(sheet: Worksheet) -> SheetSelection:
        best_match = SheetSelection(
            sheet_name=sheet.title,
            header_row_index=None,
            headers=(),
            missing_required_columns=required,
            overlap_count=0,
        )
        for row_index in range(1, min(sheet.max_row, max_header_row) + 1):
            headers = tuple(normalize_header(value) for value in _row_values(sheet, row_index))
            header_set = {value for value in headers if value}
            overlap = len(required_set & header_set)
            missing = tuple(column for column in required if column not in header_set)
            candidate = SheetSelection(
                sheet_name=sheet.title,
                header_row_index=row_index,
                headers=headers,
                missing_required_columns=missing,
                overlap_count=overlap,
            )
            if overlap > best_match.overlap_count:
                best_match = candidate
            if not missing:
                return candidate
        return best_match

    preferred_sheet = None
    if preferred_sheet_name:
        normalized_preferred = normalize_header(preferred_sheet_name).lower()
        for sheet in workbook.worksheets:
            if normalize_header(sheet.title).lower() == normalized_preferred:
                preferred_sheet = sheet
                break

    if preferred_sheet is not None:
        match = evaluate_sheet(preferred_sheet)
        if not match.missing_required_columns:
            return match

    best = SheetSelection(sheet_name=None, header_row_index=None, headers=(), missing_required_columns=required, overlap_count=0)
    for sheet in workbook.worksheets:
        candidate = evaluate_sheet(sheet)
        if candidate.overlap_count > best.overlap_count:
            best = candidate
        if not candidate.missing_required_columns:
            return candidate
    return best


def build_header_index(headers: tuple[str, ...]) -> dict[str, int]:
    return {
        header: index
        for index, header in enumerate(headers)
        if header
    }


def cell_to_decimal(cell: Cell | ReadOnlyCell | None, *, percent_mode: bool = False) -> Decimal | None:
    if cell is None:
        return None
    value = cell.value
    if value in (None, ""):
        return None
    if isinstance(value, str):
        text = value.strip()
        if not text or text == "-":
            return None
        text = text.replace("\xa0", "").replace(" ", "").replace(",", ".")
        if percent_mode:
            text = text.replace("%", "")
        try:
            return Decimal(text)
        except (ArithmeticError, InvalidOperation, ValueError):
            return None

    try:
        decimal_value = Decimal(str(value))
    except (ArithmeticError, InvalidOperation, ValueError):
        return None

    number_format = getattr(cell, "number_format", "") or ""
    if percent_mode and "%" in number_format and decimal_value <= Decimal("1"):
        return decimal_value * Decimal("100")
    return decimal_value


def cell_to_int(cell: Cell | ReadOnlyCell | None) -> int | None:
    decimal_value = cell_to_decimal(cell)
    if decimal_value is None:
        return None
    try:
        return int(decimal_value)
    except (TypeError, ValueError):
        return None


def cell_to_text(cell: Cell | ReadOnlyCell | None) -> str | None:
    if cell is None:
        return None
    return normalize_text(cell.value)


def iter_data_rows(sheet: Worksheet, header_row_index: int):
    yield from sheet.iter_rows(min_row=header_row_index + 1)


def json_ready_preview(rows: Iterable[dict[str, Any]], limit: int = 5) -> list[dict[str, Any]]:
    preview: list[dict[str, Any]] = []
    for row in rows:
        rendered: dict[str, Any] = {}
        for key, value in row.items():
            if isinstance(value, (date, datetime, Decimal)):
                rendered[key] = str(value)
            else:
                rendered[key] = value
        preview.append(rendered)
        if len(preview) >= limit:
            break
    return preview
