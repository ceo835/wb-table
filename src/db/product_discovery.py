from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import date, datetime, time, timezone
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from src.clients.wb_content_client import WBContentClient
from src.db.ad_campaign_loader import collect_ad_campaign_rows
from src.db.ad_cost_loader import collect_ad_cost_rows
from src.db.funnel_loader import collect_funnel_rows
from src.db.models import (
    FactAdCampaignNmDay,
    FactAdCostDay,
    FactAdCostEvent,
    FactFunnelDay,
    FactLocalizationRegionDay,
    FactSearchQueryMetric,
    FactStockSnapshot,
    SettingsProducts,
)
from src.db.search_query_loader import collect_search_query_rows
from src.db.session import session_scope, upsert_rows
from src.db.stock_loader import collect_stock_rows
from src.pipelines.mvp_real_run import MvpRealRun, ROOT_DIR as PIPELINE_ROOT_DIR


ROOT_DIR = PIPELINE_ROOT_DIR
DATA_DIR = ROOT_DIR / "data" / "processed"
DOCS_DIR = ROOT_DIR / "docs"
PRODUCTS_DISCOVERY_CSV_PATH = DATA_DIR / "products_discovery.csv"
PRODUCTS_DISCOVERY_MD_PATH = DOCS_DIR / "products_discovery_report.md"

DISCOVERY_START = date(2026, 5, 31)
DISCOVERY_END = date(2026, 6, 1)
DISCOVERY_SNAPSHOT_DATE = DISCOVERY_END

WORKBOOK_PATTERNS = (
    "Корзина 23.05.2026*Удалил подарки*.xlsm",
    "Корзина 23.05.2026*Удалил подарки*.xlsx",
)

FAKE_MARKERS = ("art-", "testbrand", "товар тестовый", "dry_run", "mock", "fake")
SETTINGS_PRODUCTS_CONFLICT_COLUMNS = ("nm_id",)
REPORT_COLUMNS = (
    "nm_id",
    "supplier_article",
    "title",
    "subject",
    "brand",
    "source_list",
    "first_seen_at",
    "last_seen_at",
    "data_quality_status",
    "already_in_fact_tables",
)

SOURCE_LABELS = {
    "wb_content_api": "WB Content API",
    "sales_funnel_api": "WB Sales Funnel",
    "stocks_api": "WB Stocks",
    "ad_costs_api": "WB Promotion costs",
    "fullstats_api": "WB Promotion fullstats",
    "search_queries_api": "WB Search queries",
    "excel_xlsm": "Excel/xlsm",
    "db_fact_funnel_day": "DB fact_funnel_day",
    "db_fact_stock_snapshot": "DB fact_stock_snapshot",
    "db_fact_ad_cost_event": "DB fact_ad_cost_event",
    "db_fact_ad_cost_day": "DB fact_ad_cost_day",
    "db_fact_ad_campaign_nm_day": "DB fact_ad_campaign_nm_day",
    "db_fact_search_query_metric": "DB fact_search_query_metric",
    "db_fact_localization_region_day": "DB fact_localization_region_day",
}


@dataclass(slots=True)
class SourceSummary:
    source_name: str
    status: str
    rows_observed: int
    unique_nm_ids: int
    error_short: str = ""


def _now_utc() -> datetime:
    return datetime.now(timezone.utc)


def _clean_text(value: Any) -> str | None:
    if value in (None, "", [], {}):
        return None
    text = str(value).strip()
    return text or None


def _normalize_marker_text(value: Any) -> str:
    return str(value or "").strip().lower()


def _looks_fake_text(value: Any) -> bool:
    normalized = _normalize_marker_text(value)
    return bool(normalized) and any(marker in normalized for marker in FAKE_MARKERS)


def _coerce_nm_id(value: Any) -> int | None:
    if value in (None, "", [], {}):
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value if value > 0 else None
    if isinstance(value, float):
        return int(value) if value > 0 else None
    text = str(value).strip()
    if not text:
        return None
    digits = "".join(ch for ch in text if ch.isdigit())
    if not digits:
        return None
    try:
        parsed = int(digits)
    except ValueError:
        return None
    return parsed if parsed > 0 else None


def _coerce_datetime(value: Any) -> datetime | None:
    if value in (None, "", [], {}):
        return None
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=timezone.utc)
    if isinstance(value, date):
        return datetime.combine(value, time.min, tzinfo=timezone.utc)
    text = str(value).strip()
    if not text:
        return None
    normalized = text.replace("Z", "+00:00")
    try:
        parsed = datetime.fromisoformat(normalized)
    except ValueError:
        if " " in normalized:
            normalized = normalized.replace(" ", "T", 1)
            try:
                parsed = datetime.fromisoformat(normalized)
            except ValueError:
                return None
        else:
            return None
    return parsed if parsed.tzinfo else parsed.replace(tzinfo=timezone.utc)


def _min_datetime(left: datetime | None, right: datetime | None) -> datetime | None:
    if left is None:
        return right
    if right is None:
        return left
    return left if left <= right else right


def _max_datetime(left: datetime | None, right: datetime | None) -> datetime | None:
    if left is None:
        return right
    if right is None:
        return left
    return left if left >= right else right


def _compute_quality_status(item: Mapping[str, Any]) -> str:
    fields_present = sum(
        1
        for field_name in ("supplier_article", "title", "subject", "brand")
        if _clean_text(item.get(field_name))
    )
    if fields_present == 4:
        return "COMPLETE"
    if fields_present > 0:
        return "PARTIAL"
    return "NM_ID_ONLY"


def _bool_text(value: bool) -> str:
    return "true" if value else "false"


def _format_datetime(value: datetime | None) -> str:
    if value is None:
        return ""
    return value.isoformat()


def _workbook_path() -> Path | None:
    for pattern in WORKBOOK_PATTERNS:
        matches = sorted(ROOT_DIR.glob(pattern))
        if matches:
            return matches[0]
    return None


def _header_key(value: Any) -> str:
    return "".join(ch for ch in str(value or "").strip().lower() if ch.isalnum())


def _header_matches(header_value: Any, variants: Sequence[str]) -> bool:
    normalized = _header_key(header_value)
    return any(_header_key(variant) in normalized for variant in variants)


def _fetch_all_content_cards(page_size: int = 100, max_pages: int = 100) -> tuple[list[dict[str, Any]], SourceSummary]:
    client = WBContentClient()
    error_short = ""

    try:
        cards = client.fetch_cards_catalog(limit=page_size, max_pages=max_pages, save_raw=False)
    except Exception as exc:
        error_short = str(exc)
        return [], SourceSummary("wb_content_api", "FAIL", 0, 0, error_short)

    unique_nm_ids = {
        normalized["nm_id"]
        for normalized in (client.normalize_card(card) for card in cards)
        if normalized is not None
    }
    status = "OK" if cards else "EMPTY"
    return cards, SourceSummary("wb_content_api", status, len(cards), len(unique_nm_ids), error_short)


def _content_card_to_observation(card: Mapping[str, Any], discovered_at: datetime) -> dict[str, Any] | None:
    normalized = WBContentClient.normalize_card(dict(card))
    if normalized is None:
        return None
    return _sanitize_observation(
        {
            "nm_id": normalized["nm_id"],
            "supplier_article": normalized.get("supplier_article"),
            "title": normalized.get("title"),
            "subject": normalized.get("subject"),
            "brand": normalized.get("brand"),
            "source_name": "wb_content_api",
            "seen_at": discovered_at,
        }
    )


def _row_observation(
    source_name: str,
    row: Mapping[str, Any],
    *,
    nm_keys: Sequence[str] = ("nm_id", "nmId", "nmID"),
    supplier_keys: Sequence[str] = ("supplier_article",),
    title_keys: Sequence[str] = ("title",),
    subject_keys: Sequence[str] = ("subject",),
    brand_keys: Sequence[str] = ("brand",),
    seen_keys: Sequence[str] = ("date",),
) -> dict[str, Any] | None:
    nm_id = None
    for key in nm_keys:
        nm_id = _coerce_nm_id(row.get(key))
        if nm_id is not None:
            break
    if nm_id is None:
        return None

    seen_at = None
    for key in seen_keys:
        seen_at = _coerce_datetime(row.get(key))
        if seen_at is not None:
            break

    observation = {
        "nm_id": nm_id,
        "supplier_article": next((_clean_text(row.get(key)) for key in supplier_keys if _clean_text(row.get(key))), None),
        "title": next((_clean_text(row.get(key)) for key in title_keys if _clean_text(row.get(key))), None),
        "subject": next((_clean_text(row.get(key)) for key in subject_keys if _clean_text(row.get(key))), None),
        "brand": next((_clean_text(row.get(key)) for key in brand_keys if _clean_text(row.get(key))), None),
        "source_name": source_name,
        "seen_at": seen_at or _now_utc(),
    }
    return _sanitize_observation(observation)


def _sanitize_observation(observation: Mapping[str, Any] | None) -> dict[str, Any] | None:
    if not observation:
        return None
    nm_id = _coerce_nm_id(observation.get("nm_id"))
    if nm_id is None:
        return None

    supplier_article = _clean_text(observation.get("supplier_article"))
    title = _clean_text(observation.get("title"))
    subject = _clean_text(observation.get("subject"))
    brand = _clean_text(observation.get("brand"))

    if any(_looks_fake_text(value) for value in (supplier_article, title, brand)):
        return None

    source_name = _clean_text(observation.get("source_name")) or "unknown_source"
    seen_at = _coerce_datetime(observation.get("seen_at")) or _now_utc()
    return {
        "nm_id": nm_id,
        "supplier_article": supplier_article,
        "title": title,
        "subject": subject,
        "brand": brand,
        "source_name": source_name,
        "seen_at": seen_at,
    }


def _merge_observations(
    catalog: dict[int, dict[str, Any]],
    observations: Iterable[Mapping[str, Any]],
) -> None:
    for raw_observation in observations:
        observation = _sanitize_observation(raw_observation)
        if observation is None:
            continue
        nm_id = observation["nm_id"]
        current = catalog.setdefault(
            nm_id,
            {
                "nm_id": nm_id,
                "supplier_article": None,
                "title": None,
                "subject": None,
                "brand": None,
                "source_set": set(),
                "first_seen_at": None,
                "last_seen_at": None,
            },
        )
        for field_name in ("supplier_article", "title", "subject", "brand"):
            if not current.get(field_name) and observation.get(field_name):
                current[field_name] = observation[field_name]
        current["source_set"].add(observation["source_name"])
        current["first_seen_at"] = _min_datetime(current.get("first_seen_at"), observation.get("seen_at"))
        current["last_seen_at"] = _max_datetime(current.get("last_seen_at"), observation.get("seen_at"))


def _collect_api_sources(start: date, end: date) -> tuple[dict[int, dict[str, Any]], list[SourceSummary]]:
    runner = MvpRealRun()
    runner.date_from = start
    runner.date_to = end
    catalog: dict[int, dict[str, Any]] = {}
    summaries: list[SourceSummary] = []
    discovered_at = _now_utc()

    try:
        content_cards, content_summary = _fetch_all_content_cards()
        _merge_observations(catalog, (_content_card_to_observation(card, discovered_at) for card in content_cards))
        summaries.append(content_summary)
    except Exception as exc:
        summaries.append(SourceSummary("wb_content_api", "FAIL", 0, 0, str(exc)))

    try:
        funnel_rows, funnel_meta = collect_funnel_rows(runner, start, end)
        funnel_observations = [
            _row_observation(
                "sales_funnel_api",
                row,
                seen_keys=("date",),
            )
            for row in funnel_rows
        ]
        _merge_observations(catalog, funnel_observations)
        summaries.append(
            SourceSummary(
                "sales_funnel_api",
                "OK" if funnel_rows else "EMPTY",
                len(funnel_rows),
                len({obs["nm_id"] for obs in funnel_observations if obs}),
                funnel_meta.get("history_error", "") or funnel_meta.get("products_error", ""),
            )
        )
    except Exception as exc:
        summaries.append(SourceSummary("sales_funnel_api", "FAIL", 0, 0, str(exc)))

    try:
        stock_rows, stock_meta = collect_stock_rows(runner, end)
        stock_observations = [
            _row_observation(
                "stocks_api",
                row,
                seen_keys=("snapshot_date",),
            )
            for row in stock_rows
        ]
        _merge_observations(catalog, stock_observations)
        summaries.append(
            SourceSummary(
                "stocks_api",
                "OK" if stock_rows else "EMPTY",
                len(stock_rows),
                len({obs["nm_id"] for obs in stock_observations if obs}),
                stock_meta.get("error", ""),
            )
        )
    except Exception as exc:
        summaries.append(SourceSummary("stocks_api", "FAIL", 0, 0, str(exc)))

    try:
        ad_event_rows, _ad_day_rows, ad_cost_meta = collect_ad_cost_rows(runner, start, end)
        ad_cost_observations = [
            _row_observation(
                "ad_costs_api",
                row,
                nm_keys=("nm_id", "nm_id_from_campaign_name", "nm_id_from_section"),
                seen_keys=("writeoff_datetime", "date"),
                title_keys=(),
                subject_keys=(),
                brand_keys=(),
            )
            for row in ad_event_rows
        ]
        _merge_observations(catalog, ad_cost_observations)
        summaries.append(
            SourceSummary(
                "ad_costs_api",
                "OK" if ad_event_rows else "EMPTY",
                len(ad_event_rows),
                len({obs["nm_id"] for obs in ad_cost_observations if obs}),
                ad_cost_meta.get("error", ""),
            )
        )
    except Exception as exc:
        summaries.append(SourceSummary("ad_costs_api", "FAIL", 0, 0, str(exc)))

    try:
        campaign_rows, nm_rows, fullstats_meta = collect_ad_campaign_rows(runner, start, end)
        fullstats_observations = [
            _row_observation(
                "fullstats_api",
                row,
                title_keys=("product_name",),
                subject_keys=(),
                brand_keys=(),
                seen_keys=("date",),
            )
            for row in nm_rows
        ]
        _merge_observations(catalog, fullstats_observations)
        summaries.append(
            SourceSummary(
                "fullstats_api",
                "OK" if nm_rows or campaign_rows else "EMPTY",
                len(nm_rows),
                len({obs["nm_id"] for obs in fullstats_observations if obs}),
                fullstats_meta.get("error", ""),
            )
        )
    except Exception as exc:
        summaries.append(SourceSummary("fullstats_api", "FAIL", 0, 0, str(exc)))

    try:
        search_rows, search_meta = collect_search_query_rows(runner, start, end)
        search_observations = [
            _row_observation(
                "search_queries_api",
                row,
                seen_keys=("date", "period_end"),
            )
            for row in search_rows
        ]
        _merge_observations(catalog, search_observations)
        summaries.append(
            SourceSummary(
                "search_queries_api",
                "OK" if search_rows else "EMPTY",
                len(search_rows),
                len({obs["nm_id"] for obs in search_observations if obs}),
                search_meta.get("current_status", ""),
            )
        )
    except Exception as exc:
        summaries.append(SourceSummary("search_queries_api", "FAIL", 0, 0, str(exc)))

    return catalog, summaries


def _collect_excel_source(discovered_at: datetime | None = None) -> tuple[dict[int, dict[str, Any]], SourceSummary]:
    workbook_path = _workbook_path()
    if workbook_path is None:
        return {}, SourceSummary("excel_xlsm", "SKIPPED", 0, 0, "Workbook not found in project root")

    discovered_at = discovered_at or _now_utc()
    workbook = load_workbook(workbook_path, data_only=True, keep_vba=True, read_only=True)
    catalog: dict[int, dict[str, Any]] = {}
    rows_observed = 0

    nm_variants = ("артикулwb", "артикулвб", "nmid", "nmid", "nmid", "nmid", "nmid", "nm_id", "номенклатура")
    supplier_variants = ("артикулпродавца", "артикулпоставщика", "vendorcode", "supplierarticle")
    title_variants = ("название", "товар", "наименованиетовара")
    subject_variants = ("предмет",)
    brand_variants = ("бренд",)

    for sheet in workbook.worksheets:
        sheet_max_row = sheet.max_row or 0
        if sheet_max_row <= 0:
            continue
        header_rows: list[tuple[int, dict[str, int]]] = []
        for row_index in range(1, min(sheet_max_row, 10) + 1):
            columns_map: dict[str, int] = {}
            for cell in sheet[row_index]:
                header_value = cell.value
                if header_value in (None, ""):
                    continue
                if "nm_id" not in columns_map and _header_matches(header_value, nm_variants):
                    columns_map["nm_id"] = cell.column
                if "supplier_article" not in columns_map and _header_matches(header_value, supplier_variants):
                    columns_map["supplier_article"] = cell.column
                if "title" not in columns_map and _header_matches(header_value, title_variants):
                    columns_map["title"] = cell.column
                if "subject" not in columns_map and _header_matches(header_value, subject_variants):
                    columns_map["subject"] = cell.column
                if "brand" not in columns_map and _header_matches(header_value, brand_variants):
                    columns_map["brand"] = cell.column
            if "nm_id" in columns_map:
                header_rows.append((row_index, columns_map))

        for header_row, columns_map in header_rows:
            for row_index in range(header_row + 1, sheet_max_row + 1):
                nm_id = _coerce_nm_id(sheet.cell(row_index, columns_map["nm_id"]).value)
                if nm_id is None:
                    continue
                rows_observed += 1
                observation = _sanitize_observation(
                    {
                        "nm_id": nm_id,
                        "supplier_article": sheet.cell(row_index, columns_map.get("supplier_article", 0)).value if columns_map.get("supplier_article") else None,
                        "title": sheet.cell(row_index, columns_map.get("title", 0)).value if columns_map.get("title") else None,
                        "subject": sheet.cell(row_index, columns_map.get("subject", 0)).value if columns_map.get("subject") else None,
                        "brand": sheet.cell(row_index, columns_map.get("brand", 0)).value if columns_map.get("brand") else None,
                        "source_name": "excel_xlsm",
                        "seen_at": discovered_at,
                    }
                )
                _merge_observations(catalog, [observation] if observation else [])

    workbook.close()
    status = "OK" if catalog else "EMPTY"
    return catalog, SourceSummary("excel_xlsm", status, rows_observed, len(catalog), "")


def _db_source_query(session: Session, source_name: str) -> list[dict[str, Any]]:
    if source_name == "db_fact_funnel_day":
        stmt = select(
            FactFunnelDay.nm_id,
            FactFunnelDay.date.label("seen_at"),
        )
        rows = session.execute(stmt).all()
        return [{"nm_id": row.nm_id, "seen_at": row.seen_at} for row in rows]

    if source_name == "db_fact_stock_snapshot":
        stmt = select(
            FactStockSnapshot.nm_id,
            FactStockSnapshot.supplier_article,
            FactStockSnapshot.title,
            FactStockSnapshot.subject,
            FactStockSnapshot.brand,
            FactStockSnapshot.snapshot_date.label("seen_at"),
        )
        rows = session.execute(stmt).all()
        return [row._asdict() for row in rows]

    if source_name == "db_fact_ad_cost_event":
        stmt = select(
            FactAdCostEvent.nm_id,
            FactAdCostEvent.writeoff_datetime.label("seen_at"),
        ).where(FactAdCostEvent.nm_id.is_not(None))
        rows = session.execute(stmt).all()
        return [row._asdict() for row in rows]

    if source_name == "db_fact_ad_cost_day":
        stmt = select(
            FactAdCostDay.nm_id,
            FactAdCostDay.date.label("seen_at"),
        ).where(FactAdCostDay.nm_id.is_not(None))
        rows = session.execute(stmt).all()
        return [row._asdict() for row in rows]

    if source_name == "db_fact_ad_campaign_nm_day":
        stmt = select(
            FactAdCampaignNmDay.nm_id,
            FactAdCampaignNmDay.product_name.label("title"),
            FactAdCampaignNmDay.date.label("seen_at"),
        )
        rows = session.execute(stmt).all()
        return [row._asdict() for row in rows]

    if source_name == "db_fact_search_query_metric":
        stmt = select(
            FactSearchQueryMetric.nm_id,
            FactSearchQueryMetric.supplier_article,
            FactSearchQueryMetric.title,
            FactSearchQueryMetric.subject,
            FactSearchQueryMetric.brand,
            FactSearchQueryMetric.date.label("seen_at"),
            FactSearchQueryMetric.period_end.label("period_end"),
        )
        rows = session.execute(stmt).all()
        result = []
        for row in rows:
            payload = row._asdict()
            payload["seen_at"] = payload.get("seen_at") or payload.get("period_end")
            result.append(payload)
        return result

    if source_name == "db_fact_localization_region_day":
        stmt = select(
            FactLocalizationRegionDay.nm_id,
            FactLocalizationRegionDay.supplier_article,
            FactLocalizationRegionDay.title,
            FactLocalizationRegionDay.subject,
            FactLocalizationRegionDay.brand,
            FactLocalizationRegionDay.period_end.label("seen_at"),
        )
        rows = session.execute(stmt).all()
        return [row._asdict() for row in rows]

    raise KeyError(f"Unsupported DB discovery source: {source_name}")


def _collect_db_sources() -> tuple[dict[int, dict[str, Any]], list[SourceSummary]]:
    source_names = (
        "db_fact_funnel_day",
        "db_fact_stock_snapshot",
        "db_fact_ad_cost_event",
        "db_fact_ad_cost_day",
        "db_fact_ad_campaign_nm_day",
        "db_fact_search_query_metric",
        "db_fact_localization_region_day",
    )
    catalog: dict[int, dict[str, Any]] = {}
    summaries: list[SourceSummary] = []

    with session_scope() as session:
        for source_name in source_names:
            rows = _db_source_query(session, source_name)
            observations = [
                _row_observation(
                    source_name,
                    row,
                    seen_keys=("seen_at", "period_end"),
                )
                for row in rows
            ]
            _merge_observations(catalog, observations)
            summaries.append(
                SourceSummary(
                    source_name,
                    "OK" if rows else "EMPTY",
                    len(rows),
                    len({obs["nm_id"] for obs in observations if obs}),
                    "",
                )
            )

    return catalog, summaries


def _catalog_to_rows(
    catalog: Mapping[int, Mapping[str, Any]],
    *,
    fact_table_nm_ids: set[int] | None = None,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    fact_table_nm_ids = fact_table_nm_ids or set()
    for nm_id in sorted(catalog):
        item = catalog[nm_id]
        rows.append(
            {
                "nm_id": nm_id,
                "supplier_article": item.get("supplier_article") or "",
                "title": item.get("title") or "",
                "subject": item.get("subject") or "",
                "brand": item.get("brand") or "",
                "source_list": ", ".join(sorted(item.get("source_set", set()))),
                "first_seen_at": _format_datetime(item.get("first_seen_at")),
                "last_seen_at": _format_datetime(item.get("last_seen_at")),
                "data_quality_status": _compute_quality_status(item),
                "already_in_fact_tables": _bool_text(nm_id in fact_table_nm_ids),
            }
        )
    return rows


def _load_existing_settings(session: Session) -> dict[int, SettingsProducts]:
    stmt = select(SettingsProducts)
    return {row.nm_id: row for row in session.execute(stmt).scalars()}


def _existing_source_set(settings_row: SettingsProducts | None) -> set[str]:
    if settings_row is None or not settings_row.source_list:
        return set()
    return {item.strip() for item in settings_row.source_list.split(",") if item.strip()}


def build_settings_products_upsert_rows(
    catalog: Mapping[int, Mapping[str, Any]],
    existing_rows: Mapping[int, SettingsProducts] | None = None,
) -> list[dict[str, Any]]:
    existing_rows = existing_rows or {}
    prepared_rows: list[dict[str, Any]] = []

    for nm_id in sorted(catalog):
        discovered = catalog[nm_id]
        existing = existing_rows.get(nm_id)
        source_set = set(discovered.get("source_set", set())) | _existing_source_set(existing)
        first_seen_at = discovered.get("first_seen_at")
        last_seen_at = discovered.get("last_seen_at")
        if existing is not None:
            first_seen_at = _min_datetime(existing.first_seen_at, first_seen_at)
            last_seen_at = _max_datetime(existing.last_seen_at, last_seen_at)

        prepared_rows.append(
            {
                "nm_id": nm_id,
                "supplier_article": (existing.supplier_article if existing and existing.supplier_article else discovered.get("supplier_article")) or None,
                "title": (existing.title if existing and existing.title else discovered.get("title")) or None,
                "subject": (existing.subject if existing and existing.subject else discovered.get("subject")) or None,
                "brand": (existing.brand if existing and existing.brand else discovered.get("brand")) or None,
                "active": existing.active if existing is not None else True,
                "is_new": existing.is_new if existing is not None else False,
                "report_mode": (existing.report_mode if existing and existing.report_mode else "main"),
                "group_name": existing.group_name if existing is not None else None,
                "item_type": (existing.item_type if existing and existing.item_type else "normal"),
                "source_list": ", ".join(sorted(source_set)) or None,
                "first_seen_at": first_seen_at,
                "last_seen_at": last_seen_at,
                "comment": existing.comment if existing is not None else None,
                "loaded_at": _now_utc(),
            }
        )

    return prepared_rows


def upsert_settings_products(session: Session, rows: Sequence[dict[str, Any]]) -> int:
    written = upsert_rows(
        session=session,
        model=SettingsProducts,
        rows=rows,
        conflict_columns=SETTINGS_PRODUCTS_CONFLICT_COLUMNS,
    )
    return len(rows) if written < 0 else written


def count_settings_products_rows(session: Session) -> int:
    stmt = select(func.count()).select_from(SettingsProducts)
    return int(session.execute(stmt).scalar_one())


def _write_products_csv(rows: Sequence[Mapping[str, Any]]) -> None:
    PRODUCTS_DISCOVERY_CSV_PATH.parent.mkdir(parents=True, exist_ok=True)
    with PRODUCTS_DISCOVERY_CSV_PATH.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(REPORT_COLUMNS))
        writer.writeheader()
        for row in rows:
            writer.writerow({column: row.get(column, "") for column in REPORT_COLUMNS})


def _render_report_markdown(
    *,
    total_unique_nm_ids: int,
    source_summaries: Sequence[SourceSummary],
    rows: Sequence[Mapping[str, Any]],
    fact_table_nm_ids: set[int],
    settings_rows_written: int,
    settings_rows_total: int,
    mode: str,
) -> str:
    missing_supplier = sum(1 for row in rows if not row.get("supplier_article"))
    missing_title = sum(1 for row in rows if not row.get("title"))
    sources_failed = [summary for summary in source_summaries if summary.status not in {"OK", "EMPTY"}]
    source_lines = "\n".join(
        f"| {SOURCE_LABELS.get(summary.source_name, summary.source_name)} | {summary.status} | {summary.rows_observed} | {summary.unique_nm_ids} | {summary.error_short or ''} |"
        for summary in source_summaries
    )
    failed_lines = "\n".join(
        f"- {SOURCE_LABELS.get(summary.source_name, summary.source_name)}: {summary.status} {summary.error_short}".strip()
        for summary in sources_failed
    ) or "- Нет"

    return (
        "# Products Discovery Report\n\n"
        f"- Режим: `{mode}`\n"
        f"- Период live-источников: `{DISCOVERY_START.isoformat()} .. {DISCOVERY_END.isoformat()}`\n"
        f"- Всего уникальных nm_id: **{total_unique_nm_ids}**\n"
        f"- Товаров без supplier_article: **{missing_supplier}**\n"
        f"- Товаров без title: **{missing_title}**\n"
        f"- Товаров, уже встречающихся в fact-таблицах: **{len(fact_table_nm_ids)}**\n"
        f"- Записано/обновлено в settings_products: **{settings_rows_written}**\n"
        f"- Всего строк в settings_products после запуска: **{settings_rows_total}**\n\n"
        "## Источники\n\n"
        "| Источник | Статус | Rows observed | Unique nm_id | Ошибка |\n"
        "|---|---:|---:|---:|---|\n"
        f"{source_lines}\n\n"
        "## Источники, которые не сработали\n\n"
        f"{failed_lines}\n"
    )


def _write_report_markdown(content: str) -> None:
    PRODUCTS_DISCOVERY_MD_PATH.parent.mkdir(parents=True, exist_ok=True)
    PRODUCTS_DISCOVERY_MD_PATH.write_text(content, encoding="utf-8")


def run_content_smoke(*, limit: int = 10, max_pages: int = 1) -> dict[str, Any]:
    client = WBContentClient()
    cards = client.fetch_cards_catalog(limit=limit, max_pages=max_pages, save_raw=False)
    normalized_cards = [
        normalized
        for normalized in (client.normalize_card(card) for card in cards)
        if normalized is not None
    ]
    sample = [
        {
            "nm_id": item["nm_id"],
            "supplier_article": item.get("supplier_article"),
            "title": item.get("title"),
            "brand": item.get("brand"),
            "subject": item.get("subject"),
        }
        for item in normalized_cards[:limit]
    ]

    known_match_cards = client.fetch_cards_catalog(
        limit=limit,
        max_pages=3,
        text_search="BlackWOM5",
        save_raw=False,
    )
    known_match = next(
        (
            normalized
            for normalized in (client.normalize_card(card) for card in known_match_cards)
            if normalized is not None and normalized["nm_id"] == 197330807
        ),
        None,
    )

    return {
        "cards_count": len(normalized_cards),
        "sample": sample,
        "known_nm_id_found": known_match is not None,
        "known_nm_id_card": {
            "nm_id": known_match["nm_id"],
            "supplier_article": known_match.get("supplier_article"),
            "title": known_match.get("title"),
            "brand": known_match.get("brand"),
            "subject": known_match.get("subject"),
        }
        if known_match
        else None,
    }


def discover_products(
    *,
    start: date = DISCOVERY_START,
    end: date = DISCOVERY_END,
    apply: bool = False,
) -> dict[str, Any]:
    api_catalog, api_summaries = _collect_api_sources(start, end)
    excel_catalog, excel_summary = _collect_excel_source(_now_utc())
    db_catalog, db_summaries = _collect_db_sources()

    merged_catalog: dict[int, dict[str, Any]] = {}
    for source_catalog in (api_catalog, excel_catalog, db_catalog):
        for item in source_catalog.values():
            nm_id = item["nm_id"]
            target = merged_catalog.setdefault(
                nm_id,
                {
                    "nm_id": nm_id,
                    "supplier_article": None,
                    "title": None,
                    "subject": None,
                    "brand": None,
                    "source_set": set(),
                    "first_seen_at": None,
                    "last_seen_at": None,
                },
            )
            for field_name in ("supplier_article", "title", "subject", "brand"):
                if not target.get(field_name) and item.get(field_name):
                    target[field_name] = item[field_name]
            target["source_set"].update(item.get("source_set", set()))
            target["first_seen_at"] = _min_datetime(target.get("first_seen_at"), item.get("first_seen_at"))
            target["last_seen_at"] = _max_datetime(target.get("last_seen_at"), item.get("last_seen_at"))

    fact_table_nm_ids = {nm_id for nm_id, item in db_catalog.items() if item.get("source_set")}
    rows = _catalog_to_rows(merged_catalog, fact_table_nm_ids=fact_table_nm_ids)
    _write_products_csv(rows)

    settings_rows_written = 0
    settings_rows_total = 0
    if apply:
        with session_scope() as session:
            existing_rows = _load_existing_settings(session)
            prepared_rows = build_settings_products_upsert_rows(merged_catalog, existing_rows)
            settings_rows_written = upsert_settings_products(session, prepared_rows)
            settings_rows_total = count_settings_products_rows(session)
    else:
        with session_scope() as session:
            settings_rows_total = count_settings_products_rows(session)

    source_summaries = [*api_summaries, excel_summary, *db_summaries]
    report_content = _render_report_markdown(
        total_unique_nm_ids=len(rows),
        source_summaries=source_summaries,
        rows=rows,
        fact_table_nm_ids=fact_table_nm_ids,
        settings_rows_written=settings_rows_written,
        settings_rows_total=settings_rows_total,
        mode="apply" if apply else "dry-run",
    )
    _write_report_markdown(report_content)

    return {
        "mode": "apply" if apply else "dry-run",
        "date_from": start.isoformat(),
        "date_to": end.isoformat(),
        "total_unique_nm_ids": len(rows),
        "source_counts": {summary.source_name: summary.unique_nm_ids for summary in source_summaries},
        "source_statuses": {summary.source_name: summary.status for summary in source_summaries},
        "sources_failed": [summary.source_name for summary in source_summaries if summary.status not in {"OK", "EMPTY"}],
        "without_supplier_article": sum(1 for row in rows if not row.get("supplier_article")),
        "without_title": sum(1 for row in rows if not row.get("title")),
        "already_in_fact_tables": len(fact_table_nm_ids),
        "settings_rows_written": settings_rows_written,
        "settings_rows_total": settings_rows_total,
        "duplicate_nm_ids": max(0, len(rows) - len({row["nm_id"] for row in rows})),
        "products_csv_path": str(PRODUCTS_DISCOVERY_CSV_PATH),
        "report_md_path": str(PRODUCTS_DISCOVERY_MD_PATH),
    }
