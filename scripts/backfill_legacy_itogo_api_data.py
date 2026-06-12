#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import sys
import time
from collections import defaultdict
from dataclasses import asdict, dataclass
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import load_workbook
from sqlalchemy import distinct, func, or_, select


ROOT_DIR = Path(__file__).resolve().parent.parent
SRC_DIR = ROOT_DIR / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from scripts.export_streamlit_v1_dataset import export_streamlit_v1_dataset
from src.db.ad_campaign_loader import load_ad_campaign_stats_to_db
from src.db.ad_cost_loader import load_ad_costs_to_db
from src.db.funnel_loader import load_funnel_to_db
from src.db.mart_total_report_builder import build_mart_total_report
from src.db.models import (
    FactAdCampaignDay,
    FactAdCampaignNmDay,
    FactAdCostDay,
    FactAdCostEvent,
    FactFunnelDay,
    FactSearchQueryMetric,
    FactStockSnapshot,
    SettingsProducts,
)
from src.db.search_query_loader import load_search_queries_to_db
from src.db.session import session_scope
from src.db.stock_loader import load_stocks_to_db


DEFAULT_DATE_FROM = date(2026, 5, 1)
DEFAULT_DATE_TO = date(2026, 6, 7)
DEFAULT_FUNNEL_CHUNK_SIZE = 20
DEFAULT_SEARCH_CHUNK_SIZE = 20
DEFAULT_STOCK_CHUNK_SIZE = 50
DEFAULT_FULLSTATS_SLEEP_SECONDS = 60
DEFAULT_SEARCH_SLEEP_SECONDS = 20
DEFAULT_FUNNEL_SLEEP_SECONDS = 20
DEFAULT_STOCK_SLEEP_SECONDS = 5
DATA_DIR = ROOT_DIR / "data" / "processed"
RUNTIME_SUMMARY_PATH = DATA_DIR / "legacy_itogo_api_backfill_runtime_summary.json"
FAILED_CHUNKS_PATH = DATA_DIR / "legacy_itogo_api_backfill_failed_chunks.json"


@dataclass(slots=True)
class ChunkLogRow:
    source_name: str
    report_date: str
    chunk_number: int
    item_count: int
    nm_ids: list[int]
    advert_id: int | None
    rows_fetched: int
    rows_upserted: int
    status: str
    error_type: str
    error: str
    retry_count: int
    started_at: str
    finished_at: str


def _now_utc() -> str:
    return datetime.now(timezone.utc).isoformat()


def _parse_date(value: str) -> date:
    return date.fromisoformat(value)


def _safe_fullstats_date(today: date | None = None) -> date:
    current = today or datetime.now(timezone.utc).date()
    return current - timedelta(days=2)


def _daterange(start: date, end: date) -> list[date]:
    values: list[date] = []
    current = start
    while current <= end:
        values.append(current)
        current += timedelta(days=1)
    return values


def _chunked(values: Sequence[int], size: int) -> list[list[int]]:
    if size <= 0:
        raise ValueError("chunk size must be positive")
    return [list(values[index:index + size]) for index in range(0, len(values), size)]


def _first_xlsm_path(explicit_path: Path | None = None) -> Path:
    if explicit_path is not None:
        return explicit_path
    matches = sorted(ROOT_DIR.glob("*.xlsm"))
    if not matches:
        raise FileNotFoundError("В корне проекта не найден xlsm-файл старого ИТОГО.")
    return matches[0]


def _load_legacy_itogo_nm_ids(xlsm_path: Path) -> list[int]:
    workbook = load_workbook(xlsm_path, read_only=True, data_only=True, keep_vba=True)
    worksheet = workbook[workbook.sheetnames[0]]
    seen: set[int] = set()
    nm_ids: list[int] = []
    for row in worksheet.iter_rows(min_row=3, values_only=True):
        if len(row) < 2:
            continue
        raw_nm_id = row[1]
        if not isinstance(raw_nm_id, (int, float)):
            continue
        nm_id = int(raw_nm_id)
        if nm_id in seen:
            continue
        seen.add(nm_id)
        nm_ids.append(nm_id)
    return sorted(nm_ids)


def _load_active_nm_ids() -> list[int]:
    with session_scope() as session:
        rows = session.execute(
            select(SettingsProducts.nm_id)
            .where(SettingsProducts.active.is_(True))
            .order_by(SettingsProducts.nm_id)
        ).all()
    return [int(row[0]) for row in rows if row[0] is not None]


def _load_reference_index(nm_ids: Sequence[int]) -> dict[int, dict[str, str]]:
    with session_scope() as session:
        rows = session.execute(
            select(
                SettingsProducts.nm_id,
                SettingsProducts.supplier_article,
                SettingsProducts.title,
                SettingsProducts.subject,
                SettingsProducts.brand,
            )
            .where(SettingsProducts.nm_id.in_(list(nm_ids)))
        ).all()
    return {
        int(row.nm_id): {
            "supplier_article": row.supplier_article or "",
            "title": row.title or "",
            "subject": row.subject or "",
            "brand": row.brand or "",
        }
        for row in rows
        if row.nm_id is not None
    }


def _error_type_from_text(error_text: str) -> str:
    normalized = (error_text or "").lower()
    if not normalized:
        return ""
    if "invalid start day: excess limit on days" in normalized:
        return "INVALID_DAY_LIMIT"
    if re.search(r"(^|[^0-9])429([^0-9]|$)", normalized) or "code=429" in normalized:
        return "429"
    if re.search(r"(^|[^0-9])500([^0-9]|$)", normalized) or "code=500" in normalized:
        return "500"
    if "timeout" in normalized or "timed out" in normalized:
        return "TIMEOUT"
    if "request" in normalized or "connection" in normalized:
        return "REQUEST_ERROR"
    return "FAILED_CHUNK"


def _is_retryable_error(error_type: str) -> bool:
    return error_type in {"429", "500", "TIMEOUT", "REQUEST_ERROR"}


def _status_from_result(
    source_name: str,
    result: Mapping[str, Any],
    error_text: str,
) -> tuple[str, str]:
    if error_text:
        return "FAILED_CHUNK", _error_type_from_text(error_text)

    if source_name == "search":
        if result.get("current_failed_pages") or result.get("prev_failed_pages"):
            return "FAILED_PAGES", "FAILED_PAGES"
        if int(result.get("rows_fetched", 0) or 0) == 0:
            return "NO_DATA", "NO_DATA"
        return "OK", ""

    if source_name == "stocks":
        if result.get("failed_pages"):
            return "FAILED_PAGES", "FAILED_PAGES"
        if int(result.get("rows_fetched", 0) or 0) == 0:
            return "EMPTY_RESPONSE", "EMPTY_RESPONSE"
        return "OK", ""

    if source_name == "funnel":
        if int(result.get("rows_fetched", 0) or 0) == 0:
            return "NO_DATA", "NO_DATA"
        return "OK", ""

    if source_name == "ad_cost":
        if int(result.get("event_rows_fetched", 0) or 0) == 0 and int(result.get("day_rows_built", 0) or 0) == 0:
            return "NO_DATA", "NO_DATA"
        return "OK", ""

    if source_name == "fullstats":
        if int(result.get("nm_rows_fetched", 0) or 0) == 0 and int(result.get("campaign_rows_fetched", 0) or 0) == 0:
            return "NO_DATA", "NO_DATA"
        return "OK", ""

    return "OK", ""


def _persist_runtime_state(state: Mapping[str, Any]) -> None:
    RUNTIME_SUMMARY_PATH.parent.mkdir(parents=True, exist_ok=True)
    RUNTIME_SUMMARY_PATH.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")
    FAILED_CHUNKS_PATH.write_text(
        json.dumps(state.get("failed_chunks", []), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _load_failed_chunks() -> list[dict[str, Any]]:
    if not FAILED_CHUNKS_PATH.exists():
        return []
    return json.loads(FAILED_CHUNKS_PATH.read_text(encoding="utf-8"))


def _build_failed_chunk_record(
    *,
    source_name: str,
    report_date: str,
    chunk_number: int,
    nm_ids: Sequence[int],
    advert_id: int | None,
    status: str,
    error_type: str,
    error: str,
) -> dict[str, Any]:
    return {
        "source_name": source_name,
        "report_date": report_date,
        "chunk_number": chunk_number,
        "nm_ids": list(nm_ids),
        "advert_id": advert_id,
        "status": status,
        "error_type": error_type,
        "error": error,
    }


def _distinct_nm_ids(model, column, *conditions: Any) -> set[int]:
    with session_scope() as session:
        rows = session.execute(select(distinct(column)).where(*conditions)).all()
    return {int(row[0]) for row in rows if row[0] is not None}


def _meaningful_funnel_nm_ids(start: date, end: date, nm_ids: Sequence[int]) -> set[int]:
    with session_scope() as session:
        rows = session.execute(
            select(distinct(FactFunnelDay.nm_id)).where(
                FactFunnelDay.date >= start,
                FactFunnelDay.date <= end,
                FactFunnelDay.nm_id.in_(list(nm_ids)),
                or_(
                    FactFunnelDay.card_clicks.is_not(None),
                    FactFunnelDay.cart_count.is_not(None),
                    FactFunnelDay.order_count.is_not(None),
                    FactFunnelDay.order_sum.is_not(None),
                    FactFunnelDay.add_to_cart_conversion.is_not(None),
                    FactFunnelDay.cart_to_order_conversion.is_not(None),
                    FactFunnelDay.buyout_count.is_not(None),
                    FactFunnelDay.buyout_sum.is_not(None),
                ),
            )
        ).all()
    return {int(row[0]) for row in rows if row[0] is not None}


def _source_sets(start: date, end: date, nm_ids: Sequence[int]) -> dict[str, set[int]]:
    scoped = list(nm_ids)
    return {
        "funnel": _meaningful_funnel_nm_ids(start, end, scoped),
        "search": _distinct_nm_ids(
            FactSearchQueryMetric,
            FactSearchQueryMetric.nm_id,
            FactSearchQueryMetric.period_start >= start,
            FactSearchQueryMetric.period_end <= end,
            FactSearchQueryMetric.nm_id.in_(scoped),
        ),
        "stocks": _distinct_nm_ids(
            FactStockSnapshot,
            FactStockSnapshot.nm_id,
            FactStockSnapshot.snapshot_date >= start,
            FactStockSnapshot.snapshot_date <= end,
            FactStockSnapshot.nm_id.in_(scoped),
        ),
        "ad_cost": _distinct_nm_ids(
            FactAdCostDay,
            FactAdCostDay.nm_id,
            FactAdCostDay.date >= start,
            FactAdCostDay.date <= end,
            FactAdCostDay.nm_id.in_(scoped),
        ),
        "fullstats": _distinct_nm_ids(
            FactAdCampaignNmDay,
            FactAdCampaignNmDay.nm_id,
            FactAdCampaignNmDay.date >= start,
            FactAdCampaignNmDay.date <= end,
            FactAdCampaignNmDay.nm_id.in_(scoped),
        ),
    }


def _coverage_snapshot(
    *,
    start: date,
    end: date,
    active_nm_ids: Sequence[int],
    target_nm_ids: Sequence[int],
) -> dict[str, Any]:
    active_sets = _source_sets(start, end, active_nm_ids)
    target_sets = _source_sets(start, end, target_nm_ids)

    def _scope_summary(scope_nm_ids: Sequence[int], source_sets: Mapping[str, set[int]]) -> dict[str, Any]:
        base_set = set(scope_nm_ids)
        union_any = set().union(*source_sets.values()) if source_sets else set()
        sparse_count = 0
        for nm_id in base_set:
            source_hits = sum(1 for values in source_sets.values() if nm_id in values)
            if source_hits <= 1:
                sparse_count += 1
        return {
            "products_total": len(base_set),
            "funnel_coverage": len(source_sets["funnel"]),
            "search_coverage": len(source_sets["search"]),
            "stock_coverage": len(source_sets["stocks"]),
            "ad_cost_coverage": len(source_sets["ad_cost"]),
            "ads_coverage": len(source_sets["fullstats"]),
            "with_any_metric": len(union_any),
            "without_any_metric": len(base_set - union_any),
            "almost_empty_count": sparse_count,
        }

    return {
        "active": _scope_summary(active_nm_ids, active_sets),
        "target": _scope_summary(target_nm_ids, target_sets),
    }


def _load_ad_event_groups(start: date, end: date, nm_ids: Sequence[int]) -> list[tuple[int, list[dict[str, Any]]]]:
    with session_scope() as session:
        rows = session.execute(
            select(
                FactAdCostEvent.advert_id,
                FactAdCostEvent.campaign_name,
                FactAdCostEvent.nm_id,
            ).where(
                FactAdCostEvent.date >= start,
                FactAdCostEvent.date <= end,
                FactAdCostEvent.nm_id.in_(list(nm_ids)),
                FactAdCostEvent.advert_id.is_not(None),
            )
        ).all()
    grouped: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        advert_id = int(row.advert_id)
        grouped[advert_id].append(
            {
                "advertId": advert_id,
                "campaign_name": row.campaign_name,
                "nm_id": int(row.nm_id),
            }
        )
    return [(advert_id, grouped[advert_id]) for advert_id in sorted(grouped)]


def _processed_fullstats_advert_ids(start: date, end: date) -> set[int]:
    with session_scope() as session:
        rows = session.execute(
            select(distinct(FactAdCampaignDay.advert_id)).where(
                FactAdCampaignDay.date >= start,
                FactAdCampaignDay.date <= end,
                FactAdCampaignDay.advert_id.is_not(None),
            )
        ).all()
    return {int(row[0]) for row in rows if row[0] is not None}


def _target_nm_reason_map(
    *,
    target_nm_ids: Sequence[int],
    failed_chunks: Sequence[Mapping[str, Any]],
    start: date,
    end: date,
) -> dict[str, list[str]]:
    source_sets = _source_sets(start, end, target_nm_ids)
    reasons: dict[str, list[str]] = {}
    for nm_id in target_nm_ids:
        nm_reasons: list[str] = []
        for source_name in ("search", "stocks", "funnel", "fullstats"):
            if nm_id in source_sets[source_name]:
                continue
            matching_failures = [
                failed
                for failed in failed_chunks
                if failed.get("source_name") == source_name and nm_id in {int(value) for value in failed.get("nm_ids", [])}
            ]
            if matching_failures:
                statuses = sorted({str(item.get("error_type") or item.get("status") or "FAILED_CHUNK") for item in matching_failures})
                nm_reasons.append(f"{source_name}: {', '.join(statuses)}")
            else:
                nm_reasons.append(f"{source_name}: no data")
        reasons[str(nm_id)] = nm_reasons
    return reasons


def _run_with_retry(
    *,
    source_name: str,
    report_date: str,
    chunk_number: int,
    nm_ids: Sequence[int],
    advert_id: int | None,
    loader: Any,
    rows_fetched_key: str,
    rows_upserted_key: str,
    max_retries: int,
    retry_sleep_seconds: int,
) -> tuple[dict[str, Any], ChunkLogRow, dict[str, Any] | None]:
    started_at = _now_utc()
    retry_count = 0
    error_text = ""
    result: dict[str, Any] = {}
    status = "FAILED_CHUNK"
    error_type = ""
    while True:
        try:
            result = loader()
            status, error_type = _status_from_result(source_name, result, "")
            error_text = ""
            break
        except Exception as exc:
            error_text = str(exc)
            error_type = _error_type_from_text(error_text)
            status = "FAILED_CHUNK"
            if not _is_retryable_error(error_type) or retry_count >= max_retries:
                break
            retry_count += 1
            time.sleep(retry_sleep_seconds * retry_count)

    finished_at = _now_utc()
    log_row = ChunkLogRow(
        source_name=source_name,
        report_date=report_date,
        chunk_number=chunk_number,
        item_count=len(nm_ids),
        nm_ids=list(nm_ids),
        advert_id=advert_id,
        rows_fetched=int(result.get(rows_fetched_key, 0) or 0),
        rows_upserted=int(result.get(rows_upserted_key, 0) or 0),
        status=status,
        error_type=error_type,
        error=error_text,
        retry_count=retry_count,
        started_at=started_at,
        finished_at=finished_at,
    )
    failed_record = None
    if log_row.status != "OK":
        failed_record = _build_failed_chunk_record(
            source_name=source_name,
            report_date=report_date,
            chunk_number=chunk_number,
            nm_ids=nm_ids,
            advert_id=advert_id,
            status=log_row.status,
            error_type=log_row.error_type,
            error=log_row.error,
        )
    return result, log_row, failed_record


def _build_initial_state(
    *,
    xlsm_path: Path,
    active_nm_ids: Sequence[int],
    target_nm_ids: Sequence[int],
    date_from: date,
    date_to: date,
    resume_failed_only: bool,
    before_coverage: Mapping[str, Any],
) -> dict[str, Any]:
    return {
        "started_at": _now_utc(),
        "date_from": date_from.isoformat(),
        "date_to": date_to.isoformat(),
        "xlsm_path": str(xlsm_path),
        "active_products_count": len(active_nm_ids),
        "legacy_itogo_target_count": len(target_nm_ids),
        "resume_failed_only": resume_failed_only,
        "before_coverage": before_coverage,
        "source_logs": [],
        "failed_chunks": [],
        "http_error_counts": {"429": 0, "500": 0, "TIMEOUT": 0, "REQUEST_ERROR": 0},
        "summary_by_source": {},
    }


def _append_log(
    state: dict[str, Any],
    log_row: ChunkLogRow,
    failed_record: dict[str, Any] | None,
) -> None:
    state["source_logs"].append(asdict(log_row))
    if failed_record is not None:
        state["failed_chunks"].append(failed_record)
    if log_row.error_type in state["http_error_counts"]:
        state["http_error_counts"][log_row.error_type] += 1
    summary = state["summary_by_source"].setdefault(
        log_row.source_name,
        {"runs": 0, "rows_fetched": 0, "rows_upserted": 0, "status_counts": {}},
    )
    summary["runs"] += 1
    summary["rows_fetched"] += log_row.rows_fetched
    summary["rows_upserted"] += log_row.rows_upserted
    summary["status_counts"][log_row.status] = int(summary["status_counts"].get(log_row.status, 0) or 0) + 1


def _parse_resume_chunks(
    failed_chunks: Sequence[Mapping[str, Any]],
    source_name: str,
) -> list[dict[str, Any]]:
    return [dict(item) for item in failed_chunks if item.get("source_name") == source_name]


def run_backfill(
    *,
    xlsm_path: Path,
    date_from: date,
    date_to: date,
    funnel_chunk_size: int,
    search_chunk_size: int,
    stock_chunk_size: int,
    fullstats_sleep_seconds: int,
    resume_failed_only: bool,
) -> dict[str, Any]:
    active_nm_ids = _load_active_nm_ids()
    legacy_nm_ids = _load_legacy_itogo_nm_ids(xlsm_path)
    target_nm_ids = sorted(set(active_nm_ids) & set(legacy_nm_ids))
    if not target_nm_ids:
        raise RuntimeError("Не удалось определить target nm_id из старого ИТОГО среди active товаров.")

    before_coverage = _coverage_snapshot(
        start=date_from,
        end=date_to,
        active_nm_ids=active_nm_ids,
        target_nm_ids=target_nm_ids,
    )
    failed_chunks_to_resume = _load_failed_chunks() if resume_failed_only else []
    state = _build_initial_state(
        xlsm_path=xlsm_path,
        active_nm_ids=active_nm_ids,
        target_nm_ids=target_nm_ids,
        date_from=date_from,
        date_to=date_to,
        resume_failed_only=resume_failed_only,
        before_coverage=before_coverage,
    )
    if resume_failed_only:
        state["resume_failed_chunks_loaded"] = len(failed_chunks_to_resume)
    _persist_runtime_state(state)
    reference_index = _load_reference_index(target_nm_ids)

    search_failed = _parse_resume_chunks(failed_chunks_to_resume, "search")
    stocks_failed = _parse_resume_chunks(failed_chunks_to_resume, "stocks")
    funnel_failed = _parse_resume_chunks(failed_chunks_to_resume, "funnel")
    ad_cost_failed = _parse_resume_chunks(failed_chunks_to_resume, "ad_cost")
    fullstats_failed = _parse_resume_chunks(failed_chunks_to_resume, "fullstats")

    search_plan = search_failed if resume_failed_only else [
        {
            "report_date": report_date.isoformat(),
            "chunk_number": index,
            "nm_ids": chunk,
        }
        for report_date in _daterange(date_from, date_to)
        for index, chunk in enumerate(_chunked(target_nm_ids, search_chunk_size), start=1)
    ]
    for item in search_plan:
        report_date = date.fromisoformat(str(item["report_date"]))
        start_day = report_date - timedelta(days=1)
        chunk_nm_ids = [int(value) for value in item.get("nm_ids", [])]
        result, log_row, failed_record = _run_with_retry(
            source_name="search",
            report_date=report_date.isoformat(),
            chunk_number=int(item.get("chunk_number", 1) or 1),
            nm_ids=chunk_nm_ids,
            advert_id=None,
            loader=lambda start_day=start_day, report_date=report_date, chunk_nm_ids=chunk_nm_ids: load_search_queries_to_db(
                start_day,
                report_date,
                nm_ids=chunk_nm_ids,
                reference_index={nm_id: reference_index[nm_id] for nm_id in chunk_nm_ids if nm_id in reference_index},
            ),
            rows_fetched_key="rows_fetched",
            rows_upserted_key="rows_upserted",
            max_retries=3,
            retry_sleep_seconds=20,
        )
        _append_log(state, log_row, failed_record)
        _persist_runtime_state(state)
        if failed_record is None and log_row.status == "OK":
            time.sleep(DEFAULT_SEARCH_SLEEP_SECONDS)

    stocks_plan = stocks_failed if resume_failed_only else [
        {
            "report_date": report_date.isoformat(),
            "chunk_number": index,
            "nm_ids": chunk,
        }
        for report_date in _daterange(date_from, date_to)
        for index, chunk in enumerate(_chunked(target_nm_ids, stock_chunk_size), start=1)
    ]
    for item in stocks_plan:
        report_date = date.fromisoformat(str(item["report_date"]))
        chunk_nm_ids = [int(value) for value in item.get("nm_ids", [])]
        result, log_row, failed_record = _run_with_retry(
            source_name="stocks",
            report_date=report_date.isoformat(),
            chunk_number=int(item.get("chunk_number", 1) or 1),
            nm_ids=chunk_nm_ids,
            advert_id=None,
            loader=lambda report_date=report_date, chunk_nm_ids=chunk_nm_ids: load_stocks_to_db(report_date, nm_ids=chunk_nm_ids),
            rows_fetched_key="rows_fetched",
            rows_upserted_key="rows_upserted",
            max_retries=3,
            retry_sleep_seconds=15,
        )
        _append_log(state, log_row, failed_record)
        _persist_runtime_state(state)
        if failed_record is None and log_row.status == "OK":
            time.sleep(DEFAULT_STOCK_SLEEP_SECONDS)

    funnel_plan = funnel_failed if resume_failed_only else [
        {
            "report_date": report_date.isoformat(),
            "chunk_number": index,
            "nm_ids": chunk,
        }
        for report_date in _daterange(date_from, date_to)
        for index, chunk in enumerate(_chunked(target_nm_ids, funnel_chunk_size), start=1)
    ]
    for item in funnel_plan:
        report_date = date.fromisoformat(str(item["report_date"]))
        chunk_nm_ids = [int(value) for value in item.get("nm_ids", [])]
        result, log_row, failed_record = _run_with_retry(
            source_name="funnel",
            report_date=report_date.isoformat(),
            chunk_number=int(item.get("chunk_number", 1) or 1),
            nm_ids=chunk_nm_ids,
            advert_id=None,
            loader=lambda report_date=report_date, chunk_nm_ids=chunk_nm_ids: load_funnel_to_db(report_date, report_date, nm_ids=chunk_nm_ids),
            rows_fetched_key="rows_fetched",
            rows_upserted_key="rows_upserted",
            max_retries=3,
            retry_sleep_seconds=20,
        )
        _append_log(state, log_row, failed_record)
        _persist_runtime_state(state)
        if failed_record is None and log_row.status == "OK":
            time.sleep(DEFAULT_FUNNEL_SLEEP_SECONDS)

    ad_cost_plan = ad_cost_failed if resume_failed_only else [
        {
            "report_date": report_date.isoformat(),
            "chunk_number": 1,
            "nm_ids": target_nm_ids,
        }
        for report_date in _daterange(date_from, date_to)
    ]
    for item in ad_cost_plan:
        report_date = date.fromisoformat(str(item["report_date"]))
        chunk_nm_ids = [int(value) for value in item.get("nm_ids", [])]
        result, log_row, failed_record = _run_with_retry(
            source_name="ad_cost",
            report_date=report_date.isoformat(),
            chunk_number=1,
            nm_ids=chunk_nm_ids,
            advert_id=None,
            loader=lambda report_date=report_date, chunk_nm_ids=chunk_nm_ids: load_ad_costs_to_db(report_date, report_date, nm_ids=chunk_nm_ids),
            rows_fetched_key="event_rows_fetched",
            rows_upserted_key="event_rows_upserted",
            max_retries=3,
            retry_sleep_seconds=15,
        )
        _append_log(state, log_row, failed_record)
        _persist_runtime_state(state)

    processed_before = _processed_fullstats_advert_ids(date_from, date_to)
    safe_fullstats_end = min(date_to, _safe_fullstats_date())
    ad_event_groups = _load_ad_event_groups(date_from, safe_fullstats_end, target_nm_ids) if safe_fullstats_end >= date_from else []
    fullstats_plan = fullstats_failed if resume_failed_only else [
        {
            "report_date": f"{date_from.isoformat()}..{safe_fullstats_end.isoformat()}",
            "chunk_number": index,
            "advert_id": advert_id,
            "nm_ids": sorted({int(row['nm_id']) for row in rows if row.get('nm_id') is not None}),
        }
        for index, (advert_id, rows) in enumerate(ad_event_groups, start=1)
        if advert_id not in processed_before
    ]
    grouped_index = {advert_id: rows for advert_id, rows in ad_event_groups}
    for item in fullstats_plan:
        advert_id = int(item["advert_id"])
        ad_event_rows = grouped_index.get(advert_id, [])
        chunk_nm_ids = [int(value) for value in item.get("nm_ids", [])]
        result, log_row, failed_record = _run_with_retry(
            source_name="fullstats",
            report_date=str(item["report_date"]),
            chunk_number=int(item.get("chunk_number", 1) or 1),
            nm_ids=chunk_nm_ids,
            advert_id=advert_id,
            loader=lambda ad_event_rows=ad_event_rows, chunk_nm_ids=chunk_nm_ids: load_ad_campaign_stats_to_db(
                date_from,
                safe_fullstats_end,
                nm_ids=chunk_nm_ids,
                ad_event_rows=ad_event_rows,
            ),
            rows_fetched_key="nm_rows_fetched",
            rows_upserted_key="nm_rows_upserted",
            max_retries=1,
            retry_sleep_seconds=20,
        )
        _append_log(state, log_row, failed_record)
        _persist_runtime_state(state)
        if failed_record is None and log_row.status == "OK":
            time.sleep(fullstats_sleep_seconds)
        if log_row.error_type == "429":
            break

    mart_summary = build_mart_total_report(date_from, date_to, version="v2")
    streamlit_summary = export_streamlit_v1_dataset(date_from, date_to)
    after_coverage = _coverage_snapshot(
        start=date_from,
        end=date_to,
        active_nm_ids=active_nm_ids,
        target_nm_ids=target_nm_ids,
    )
    state["mart_summary"] = mart_summary
    state["streamlit_dataset_summary"] = streamlit_summary
    state["after_coverage"] = after_coverage
    state["safe_fullstats_end"] = safe_fullstats_end.isoformat() if safe_fullstats_end >= date_from else None
    state["target_nm_ids"] = target_nm_ids
    state["target_nm_ids_not_covered_reasons"] = _target_nm_reason_map(
        target_nm_ids=target_nm_ids,
        failed_chunks=state["failed_chunks"],
        start=date_from,
        end=date_to,
    )
    state["finished_at"] = _now_utc()
    _persist_runtime_state(state)
    return state


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Long-run backfill for legacy Itogo target products.")
    parser.add_argument("--date-from", type=_parse_date, default=DEFAULT_DATE_FROM)
    parser.add_argument("--date-to", type=_parse_date, default=DEFAULT_DATE_TO)
    parser.add_argument("--funnel-chunk-size", type=int, default=DEFAULT_FUNNEL_CHUNK_SIZE)
    parser.add_argument("--search-chunk-size", type=int, default=DEFAULT_SEARCH_CHUNK_SIZE)
    parser.add_argument("--stock-chunk-size", type=int, default=DEFAULT_STOCK_CHUNK_SIZE)
    parser.add_argument("--fullstats-sleep-seconds", type=int, default=DEFAULT_FULLSTATS_SLEEP_SECONDS)
    parser.add_argument("--resume-failed-only", action="store_true")
    parser.add_argument("--xlsm-path", type=Path)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if args.date_from > args.date_to:
        raise SystemExit("--date-from must be <= --date-to")
    xlsm_path = _first_xlsm_path(args.xlsm_path)
    summary = run_backfill(
        xlsm_path=xlsm_path,
        date_from=args.date_from,
        date_to=args.date_to,
        funnel_chunk_size=args.funnel_chunk_size,
        search_chunk_size=args.search_chunk_size,
        stock_chunk_size=args.stock_chunk_size,
        fullstats_sleep_seconds=args.fullstats_sleep_seconds,
        resume_failed_only=args.resume_failed_only,
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 1 if summary.get("failed_chunks") else 0


if __name__ == "__main__":
    raise SystemExit(main())
