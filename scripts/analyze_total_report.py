#!/usr/bin/env python3
from __future__ import annotations

import csv
import json
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.formula import Tokenizer
from openpyxl.utils import get_column_letter


ROOT_DIR = Path(__file__).resolve().parent.parent
DOCS_DIR = ROOT_DIR / "docs"
DATA_DIR = ROOT_DIR / "data" / "processed"
SPEC_PATH = DOCS_DIR / "TOTAL_REPORT_SPEC.md"
FORMULA_MAP_PATH = DOCS_DIR / "TOTAL_REPORT_FORMULA_MAP.md"
CSV_PATH = DATA_DIR / "total_report_column_map.csv"

TOTAL_SHEET_NAME = "Итого"
HEADER_ROW = 1
FIRST_DATA_ROW = 3
WORKBOOK_PATTERNS = [
    "Корзина 23.05.2026*Удалил подарки*.xlsm",
    "Корзина 23.05.2026*Удалил подарки*.xlsx",
]

RANGE_TOKEN_PATTERN = re.compile(r"(?:'([^']+)'|([A-Za-zА-Яа-я0-9_ ]+))!")
FIXED_RANGE_PATTERN = re.compile(r"\$[A-Z]+\$?\d+")
EXTERNAL_REFERENCE_PATTERN = re.compile(r"\[[^\]]+\]")

TRANSLIT_TABLE = str.maketrans(
    {
        "а": "a",
        "б": "b",
        "в": "v",
        "г": "g",
        "д": "d",
        "е": "e",
        "ё": "e",
        "ж": "zh",
        "з": "z",
        "и": "i",
        "й": "y",
        "к": "k",
        "л": "l",
        "м": "m",
        "н": "n",
        "о": "o",
        "п": "p",
        "р": "r",
        "с": "s",
        "т": "t",
        "у": "u",
        "ф": "f",
        "х": "h",
        "ц": "ts",
        "ч": "ch",
        "ш": "sh",
        "щ": "sch",
        "ъ": "",
        "ы": "y",
        "ь": "",
        "э": "e",
        "ю": "yu",
        "я": "ya",
    }
)


@dataclass(frozen=True)
class SheetDescriptor:
    title: str
    target_db_table: str
    automation_status: str
    notes: str


@dataclass(frozen=True)
class BlockSpec:
    start: int
    end: int
    name: str
    default_sheet: str
    target_db_table: str
    automation_status: str
    notes: str

    def contains(self, column_index: int) -> bool:
        return self.start <= column_index <= self.end


@dataclass
class SheetInfo:
    title: str
    max_row: int
    max_column: int
    formula_cells: int
    state: str


KNOWN_SOURCE_SHEETS = {
    "Воронка на день": SheetDescriptor(
        title="Воронка на день",
        target_db_table="fact_funnel_day",
        automation_status="API",
        notes="Подтягивается из WB Analytics / funnel API и уже покрывается текущим pipeline.",
    ),
    "РасходРК": SheetDescriptor(
        title="РасходРК",
        target_db_table="fact_ad_cost_event",
        automation_status="API",
        notes="Подтягивается из WB Promotion costs; часть KPI на листе Итого считается поверх расходов.",
    ),
    "РК стата": SheetDescriptor(
        title="РК стата",
        target_db_table="fact_ad_campaign_nm_day",
        automation_status="API",
        notes="Источник рекламной статистики; в боевом режиме нужен лаг минимум D-2.",
    ),
    "ВБро": SheetDescriptor(
        title="ВБро",
        target_db_table="fact_profit_day",
        automation_status="MANUAL",
        notes="Подтверждено как manual upload из внешнего сервиса; не считается через COGS/API на текущем этапе.",
    ),
    "Точка вх": SheetDescriptor(
        title="Точка вх",
        target_db_table="fact_entry_point_day",
        automation_status="MANUAL",
        notes="Нужен CSV/Excel export или private endpoint customer-profile; текущий API-источник не подтвержден.",
    ),
    "Локализация": SheetDescriptor(
        title="Локализация",
        target_db_table="fact_localization_region_day",
        automation_status="NEEDS_CONFIRMATION",
        notes="Частично покрывается region-sale, целевой источник — orders-geography export/access.",
    ),
    "Сравнение карточек": SheetDescriptor(
        title="Сравнение карточек",
        target_db_table="fact_card_comparison_metric",
        automation_status="EXTERNAL",
        notes="Зависит от внешнего comparison / MPStat блока, полностью API не закрыт.",
    ),
    "Поисковые запросы": SheetDescriptor(
        title="Поисковые запросы",
        target_db_table="fact_search_query_metric",
        automation_status="API",
        notes="Частично покрывается WB/Jam API; конкурентные перцентили и price bounds требуют отдельного подтверждения.",
    ),
}


BLOCK_SPECS = [
    BlockSpec(
        start=1,
        end=20,
        name="Базовая воронка и KPI",
        default_sheet="Воронка на день",
        target_db_table="mart_total_report",
        automation_status="API",
        notes="Базовые метрики продаж и локальные KPI по воронке.",
    ),
    BlockSpec(
        start=22,
        end=60,
        name="РасходРК и эффективность корзин",
        default_sheet="РасходРК",
        target_db_table="mart_total_report",
        automation_status="API",
        notes="Расходы, CPM, корзины от рекламы и производные показатели.",
    ),
    BlockSpec(
        start=61,
        end=63,
        name="ВБро",
        default_sheet="ВБро",
        target_db_table="mart_total_report",
        automation_status="MANUAL",
        notes="Внешний ручной сервис; в xlsm на лист Итого выведены только ссылки на ВБро.",
    ),
    BlockSpec(
        start=64,
        end=91,
        name="Точка входа: каталоги и поверхности",
        default_sheet="Точка вх",
        target_db_table="fact_entry_point_day",
        automation_status="MANUAL",
        notes="Широкий pivot по entry point; нужен export или private endpoint.",
    ),
    BlockSpec(
        start=93,
        end=107,
        name="Сравнение карточек / MPStat",
        default_sheet="Сравнение карточек",
        target_db_table="fact_card_comparison_metric",
        automation_status="EXTERNAL",
        notes="Рейтинги, средняя позиция и сопоставимые метрики карточек.",
    ),
    BlockSpec(
        start=109,
        end=121,
        name="Локализация по регионам",
        default_sheet="Локализация",
        target_db_table="fact_localization_region_day",
        automation_status="NEEDS_CONFIRMATION",
        notes="Региональные разбивки, которые в xlsm сведены в wide-формат.",
    ),
    BlockSpec(
        start=123,
        end=277,
        name="РК стата / матрица кампаний",
        default_sheet="РК стата",
        target_db_table="fact_ad_campaign_nm_day",
        automation_status="API",
        notes="Широкий динамический блок по кампаниям, который нужно нормализовать в строки.",
    ),
    BlockSpec(
        start=278,
        end=327,
        name="Портрет покупателя / точка входа",
        default_sheet="Точка вх",
        target_db_table="fact_entry_point_day",
        automation_status="MANUAL",
        notes="Разделы customer-profile и точки входа; нужен export sample или доступ в кабинет.",
    ),
    BlockSpec(
        start=333,
        end=339,
        name="Локализация summary",
        default_sheet="Локализация",
        target_db_table="fact_localization_region_summary_day",
        automation_status="NEEDS_CONFIRMATION",
        notes="Итоговые KPI локальности и региона.",
    ),
    BlockSpec(
        start=340,
        end=4000,
        name="Поисковые запросы",
        default_sheet="Поисковые запросы",
        target_db_table="fact_search_query_metric",
        automation_status="API",
        notes="Динамический wide-блок поисковых запросов; в БД должен стать row-based fact.",
    ),
]


DIRECT_FIELD_MAP = {
    "артикул продавца": ("dim_product", "supplier_article"),
    "артикул wb": ("dim_product", "nm_id"),
    "дата": ("fact_funnel_day", "date"),
    "показы": ("fact_funnel_day", "impressions"),
    "переходы в карточку": ("fact_funnel_day", "card_clicks"),
    "положили в корзину": ("fact_funnel_day", "add_to_cart"),
    "заказали, шт": ("fact_funnel_day", "orders_qty"),
    "выкупили, шт": ("fact_funnel_day", "buyout_qty"),
    "ctr": ("fact_funnel_day", "ctr"),
    "ситиар": ("mart_total_report", "site_ctr"),
    "конверсия в корзину, %": ("fact_funnel_day", "add_to_cart_conversion"),
    "конверсия в заказ, %": ("fact_funnel_day", "cart_to_order_conversion"),
    "заказали на сумму, ₽": ("fact_funnel_day", "order_sum"),
    "локальные заказы, %": ("fact_localization_region_day", "local_orders_percent"),
    "не локальные заказы, %": ("fact_localization_region_day", "nonlocal_orders_percent"),
    "время доставки": ("fact_localization_region_day", "delivery_time"),
    "сумма кампания": ("fact_ad_cost_day", "ad_spend"),
    "затраты рк": ("fact_ad_cost_event", "amount"),
    "средняя позиция": ("fact_card_comparison_metric", "avg_position"),
    "рейтинг карточки": ("fact_card_comparison_metric", "card_rating"),
    "рейтинг по отзывам": ("fact_card_comparison_metric", "review_rating"),
    "поисковые запросы": ("fact_search_query_metric", "search_query"),
    "регион": ("fact_localization_region_summary_day", "region"),
}


def resolve_workbook_path() -> Path:
    for pattern in WORKBOOK_PATTERNS:
        matches = sorted(ROOT_DIR.glob(pattern))
        if matches:
            return matches[0]
    raise FileNotFoundError("Workbook not found in project root by expected patterns.")


def markdown_escape(value: object) -> str:
    return str(value).replace("|", "\\|")


def slugify(text: str, fallback: str) -> str:
    normalized = str(text).strip().lower().translate(TRANSLIT_TABLE)
    normalized = re.sub(r"[^a-z0-9]+", "_", normalized)
    normalized = re.sub(r"_+", "_", normalized).strip("_")
    return normalized or fallback


def normalize_header(value: object, fallback: str) -> str:
    text = str(value).strip() if value not in (None, "") else ""
    return text or fallback


def find_block(column_index: int) -> BlockSpec | None:
    for spec in BLOCK_SPECS:
        if spec.contains(column_index):
            return spec
    return None


def load_workbook_info(path: Path) -> tuple[object, list[SheetInfo], dict[str, object]]:
    workbook = load_workbook(path, data_only=False, keep_vba=True)
    sheet_infos: list[SheetInfo] = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        formula_cells = 0
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_cells += 1
        sheet_infos.append(
            SheetInfo(
                title=sheet_name,
                max_row=sheet.max_row,
                max_column=sheet.max_column,
                formula_cells=formula_cells,
                state=getattr(sheet, "sheet_state", "visible"),
            )
        )

    with ZipFile(path) as archive:
        names = archive.namelist()
        external_links = [name for name in names if name.startswith("xl/externalLinks/")]
        zip_info = {
            "has_vba": "xl/vbaProject.bin" in names,
            "external_links": external_links,
        }

    return workbook, sheet_infos, zip_info


def used_column_indexes(ws) -> list[int]:
    used: list[int] = []
    for column_index in range(1, ws.max_column + 1):
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row, column_index).value
            if value not in (None, ""):
                used.append(column_index)
                break
    return used


def first_formula_in_column(ws, column_index: int) -> str:
    for row in range(FIRST_DATA_ROW, ws.max_row + 1):
        value = ws.cell(row, column_index).value
        if isinstance(value, str) and value.startswith("="):
            return value
    return ""


def formula_tokens(formula: str) -> list[str]:
    if not formula:
        return []
    try:
        tokenizer = Tokenizer(formula)
    except Exception:
        return []
    return [token.value for token in tokenizer.items if token.subtype == "RANGE"]


def parse_source_sheets(formula: str, valid_sheet_names: set[str]) -> list[str]:
    if not formula:
        return []
    sheets: list[str] = []
    for match in RANGE_TOKEN_PATTERN.finditer(formula):
        candidate = (match.group(1) or match.group(2) or "").strip()
        if candidate and candidate in valid_sheet_names and candidate != TOTAL_SHEET_NAME:
            sheets.append(candidate)
    return sorted(set(sheets))


def detect_source_type(formula: str, source_sheets: list[str]) -> str:
    if not formula:
        return "MANUAL_OR_RAW_VALUE"
    if EXTERNAL_REFERENCE_PATTERN.search(formula):
        return "EXTERNAL_LINK"
    if source_sheets:
        return "CROSS_SHEET_FORMULA"
    return "IN_SHEET_FORMULA"


def pick_status_and_notes(
    header: str,
    block: BlockSpec | None,
    source_sheets: list[str],
    source_type: str,
    has_ref_error: bool,
) -> tuple[str, str]:
    if source_type == "EXTERNAL_LINK":
        return "EXTERNAL", "В формуле есть внешняя ссылка на другой workbook."

    for source_sheet in source_sheets:
        descriptor = KNOWN_SOURCE_SHEETS.get(source_sheet)
        if descriptor:
            notes = descriptor.notes
            if has_ref_error:
                notes = f"{notes} В sample formula есть #REF!."
            return descriptor.automation_status, notes

    lowered = header.lower()
    if source_type == "IN_SHEET_FORMULA":
        notes = "Локальная формула на листе Итого; переносить в mart_total_report как вычисляемое поле."
        if has_ref_error:
            notes = f"{notes} В sample formula есть #REF!."
        return "CALCULATED", notes

    if block:
        notes = block.notes
        if has_ref_error:
            notes = f"{notes} В sample formula есть #REF!."
        return block.automation_status, notes

    if "операц" in lowered or "вбро" in lowered:
        return "MANUAL", KNOWN_SOURCE_SHEETS["ВБро"].notes
    if "точк" in lowered or "профиль покупателя" in lowered:
        return "MANUAL", KNOWN_SOURCE_SHEETS["Точка вх"].notes
    if "локал" in lowered or "регион" in lowered:
        return "NEEDS_CONFIRMATION", KNOWN_SOURCE_SHEETS["Локализация"].notes
    if "поиск" in lowered:
        return "API", KNOWN_SOURCE_SHEETS["Поисковые запросы"].notes

    return "UNKNOWN", "Нужна ручная расшифровка происхождения поля или проверка макросной логики."


def pick_target_mapping(
    header: str,
    block: BlockSpec | None,
    source_sheets: list[str],
    source_type: str,
    column_index: int,
) -> tuple[str, str]:
    normalized = " ".join(header.lower().split())
    if normalized in DIRECT_FIELD_MAP:
        return DIRECT_FIELD_MAP[normalized]

    if block and source_type == "IN_SHEET_FORMULA":
        return "mart_total_report", slugify(header, f"column_{column_index}")

    if block:
        if block.name == "Поисковые запросы" and column_index >= 341:
            return "fact_search_query_metric", "search_query"
        if block.name == "РК стата / матрица кампаний" and column_index >= 124:
            return "fact_ad_campaign_nm_day", "campaign_name"
        if block.name == "Точка входа: каталоги и поверхности":
            return "fact_entry_point_day", "entry_point"
        if block.name == "Портрет покупателя / точка входа":
            return "fact_entry_point_day", "section"
        if block.name == "Локализация по регионам":
            return "fact_localization_region_day", "region"
        if block.name == "Локализация summary":
            return "fact_localization_region_summary_day", slugify(header, f"column_{column_index}")
        if block.name == "Сравнение карточек / MPStat":
            return "fact_card_comparison_metric", slugify(header, f"column_{column_index}")
        return block.target_db_table, slugify(header, f"column_{column_index}")

    for source_sheet in source_sheets:
        descriptor = KNOWN_SOURCE_SHEETS.get(source_sheet)
        if descriptor:
            return descriptor.target_db_table, slugify(header, f"column_{column_index}")

    return "mart_total_report", slugify(header, f"column_{column_index}")


def first_non_empty_sample(ws, column_index: int) -> object:
    for row in range(FIRST_DATA_ROW, min(ws.max_row, FIRST_DATA_ROW + 25) + 1):
        value = ws.cell(row, column_index).value
        if value not in (None, ""):
            return value
    return ""


def summarize_formula_pattern(column_name: str, formula: str, source_sheets: list[str]) -> str:
    if not formula:
        return "Колонка без формулы, заполняется raw/manual или через pivot headers."
    compact = formula.replace(" ", "")
    if "VLOOKUP(" in formula and source_sheets:
        return f"Точечный lookup в лист `{source_sheets[0]}` по nm_id/заголовку с жестко зафиксированным диапазоном."
    if "SUMIF(" in formula or "SUMIFS(" in formula:
        if source_sheets:
            return f"Агрегация через SUMIF/SUMIFS по данным листа `{source_sheets[0]}` с фиксированными диапазонами."
        return "Агрегация через SUMIF/SUMIFS; диапазоны зафиксированы и хрупки к изменению структуры."
    if "MATCH(" in formula and "VLOOKUP(" in formula:
        return "Двухшаговый lookup: поиск колонки через MATCH и выбор значения через VLOOKUP."
    if "/" in compact and "*" not in compact:
        return "Локальный расчет отношения двух метрик на листе Итого."
    if "*1000" in compact:
        return "Локальный CPM/стоимость за 1000 показов на основе расходов и показов."
    if "IFERROR(" in formula:
        return "Локальная формула с подавлением ошибок; пустые/нулевые значения скрывают пропуски источника."
    if "#REF!" in formula:
        return "Формула повреждена: в образце уже есть #REF!."
    return "Локальная бизнес-формула, которую нужно перенести в mart_total_report после явной спецификации."


def analyze_total_sheet(workbook) -> tuple[list[dict[str, object]], dict[str, object]]:
    valid_sheet_names = set(workbook.sheetnames)
    ws = workbook[TOTAL_SHEET_NAME]
    used_columns = used_column_indexes(ws)
    source_sheet_counter: Counter[str] = Counter()
    formula_columns = 0
    cross_sheet_columns = 0
    external_columns = 0
    fixed_range_columns = 0
    ref_error_columns = 0
    columns: list[dict[str, object]] = []
    ref_error_samples: list[dict[str, object]] = []

    for column_index in used_columns:
        header = normalize_header(ws.cell(HEADER_ROW, column_index).value, f"UNNAMED_{get_column_letter(column_index)}")
        formula = first_formula_in_column(ws, column_index)
        tokens = formula_tokens(formula)
        source_sheets = parse_source_sheets(formula, valid_sheet_names)
        source_type = detect_source_type(formula, source_sheets)
        block = find_block(column_index)
        has_fixed_range = bool(formula and FIXED_RANGE_PATTERN.search(formula))
        has_ref_error = "#REF!" in formula if formula else False
        sample_value = first_non_empty_sample(ws, column_index)

        if formula:
            formula_columns += 1
        if source_sheets:
            cross_sheet_columns += 1
            for sheet_name in source_sheets:
                source_sheet_counter[sheet_name] += 1
        if source_type == "EXTERNAL_LINK":
            external_columns += 1
        if has_fixed_range:
            fixed_range_columns += 1
        if has_ref_error:
            ref_error_columns += 1
            if len(ref_error_samples) < 12:
                ref_error_samples.append(
                    {
                        "column_index": column_index,
                        "column_name": header,
                        "formula_example": formula,
                    }
                )

        automation_status, notes = pick_status_and_notes(
            header=header,
            block=block,
            source_sheets=source_sheets,
            source_type=source_type,
            has_ref_error=has_ref_error,
        )
        target_db_table, target_db_field = pick_target_mapping(
            header=header,
            block=block,
            source_sheets=source_sheets,
            source_type=source_type,
            column_index=column_index,
        )

        columns.append(
            {
                "column_index": column_index,
                "column_letter": get_column_letter(column_index),
                "column_name": header,
                "block_name": block.name if block else "Прочее / требует ручной группировки",
                "formula_example": formula,
                "source_sheet": ", ".join(source_sheets) if source_sheets else (block.default_sheet if block else ""),
                "source_range_or_reference": ", ".join(tokens[:8]),
                "source_type": source_type,
                "target_db_table": target_db_table,
                "target_db_field": target_db_field,
                "automation_status": automation_status,
                "notes": notes,
                "sample_value": sample_value,
                "has_ref_error": has_ref_error,
                "has_fixed_range": has_fixed_range,
                "needs_confirmation": "YES" if automation_status in {"UNKNOWN", "NEEDS_CONFIRMATION"} or has_ref_error else "NO",
                "formula_pattern": summarize_formula_pattern(header, formula, source_sheets),
            }
        )

    summary = {
        "sheet_name": TOTAL_SHEET_NAME,
        "used_column_count": len(used_columns),
        "last_nonempty_column_index": max(used_columns),
        "max_row": ws.max_row,
        "formula_columns": formula_columns,
        "cross_sheet_columns": cross_sheet_columns,
        "external_columns": external_columns,
        "fixed_range_columns": fixed_range_columns,
        "ref_error_columns": ref_error_columns,
        "source_sheets": sorted(source_sheet_counter.keys()),
        "source_sheet_usage": dict(source_sheet_counter),
        "ref_error_samples": ref_error_samples,
        "merged_ranges": [str(rng) for rng in ws.merged_cells.ranges],
        "frozen_panes": str(ws.freeze_panes) if ws.freeze_panes else "",
        "auto_filter": str(ws.auto_filter.ref) if ws.auto_filter and ws.auto_filter.ref else "",
        "block_usage": dict(Counter(column["block_name"] for column in columns)),
        "status_usage": dict(Counter(column["automation_status"] for column in columns)),
    }
    return columns, summary


def build_workbook_structure_table(sheet_infos: list[SheetInfo]) -> list[str]:
    lines = [
        "## Структура workbook",
        "",
        "| Sheet | Rows | Cols | Formula cells | State |",
        "| --- | ---: | ---: | ---: | --- |",
    ]
    for sheet in sheet_infos:
        lines.append(f"| {sheet.title} | {sheet.max_row} | {sheet.max_column} | {sheet.formula_cells} | {sheet.state} |")
    return lines


def build_source_dependencies_section(summary: dict[str, object]) -> list[str]:
    lines = [
        "## Зависимости листа Итого",
        "",
        "| Source sheet | Role in Итого | Current automation status | Notes | Referenced columns |",
        "| --- | --- | --- | --- | ---: |",
    ]
    for source_sheet in summary["source_sheets"]:
        descriptor = KNOWN_SOURCE_SHEETS.get(source_sheet)
        if descriptor:
            lines.append(
                f"| {source_sheet} | источник raw/pivot данных | {descriptor.automation_status} | {descriptor.notes} | {summary['source_sheet_usage'][source_sheet]} |"
            )
    return lines


def build_block_summary_section(summary: dict[str, object]) -> list[str]:
    lines = [
        "## Блоки колонок на Итого",
        "",
        "| Block | Column count | Comment |",
        "| --- | ---: | --- |",
    ]
    comments = {spec.name: spec.notes for spec in BLOCK_SPECS}
    for block_name, count in summary["block_usage"].items():
        lines.append(f"| {block_name} | {count} | {comments.get(block_name, 'Требует ручной группировки.')} |")
    return lines


def build_risk_section(summary: dict[str, object], zip_info: dict[str, object]) -> list[str]:
    ref_samples = summary["ref_error_samples"]
    lines = [
        "## Риски и технические особенности",
        "",
        f"- VBA/макросы: `{'yes' if zip_info['has_vba'] else 'no'}`.",
        f"- External links в package: `{len(zip_info['external_links'])}`.",
        f"- Freeze panes на `Итого`: `{summary['frozen_panes'] or 'none'}`.",
        f"- Merge ranges на `Итого`: `{', '.join(summary['merged_ranges']) or 'none'}`.",
        f"- Автофильтр на `Итого`: `{summary['auto_filter'] or 'none'}`.",
        f"- Колонки с формулами: `{summary['formula_columns']}`.",
        f"- Колонки с ссылками на другие листы: `{summary['cross_sheet_columns']}`.",
        f"- Колонки с фиксированными диапазонами: `{summary['fixed_range_columns']}`.",
        f"- Колонки, где sample formula уже содержит `#REF!`: `{summary['ref_error_columns']}`.",
        "",
        "### Что это значит для database-first переноса",
        "",
        "- В xlsm много жестких `VLOOKUP`, `SUMIF/SUMIFS` и абсолютных диапазонов: такие формулы хрупки при изменении числа артикулов, дат или кампаний.",
        "- Динамические wide-блоки `РК стата` и `Поисковые запросы` в БД надо хранить не как новые колонки, а как строковые факты (`row-based fact tables`).",
        "- Блок `Точка вх` и часть `Локализация` зависят от private/export источников, а не от подтвержденного API.",
        "- `ВБро` в текущей модели не должно пересчитываться автоматически: это manual/external сервис.",
    ]

    if ref_samples:
        lines.extend(
            [
                "",
                "### Примеры колонок с `#REF!`",
                "",
                "| column_index | column_name | formula_example |",
                "| ---: | --- | --- |",
            ]
        )
        for sample in ref_samples:
            lines.append(
                f"| {sample['column_index']} | {markdown_escape(sample['column_name'])} | `{markdown_escape(sample['formula_example'])}` |"
            )
    return lines


def build_column_table(columns: list[dict[str, object]]) -> list[str]:
    lines = [
        "## Карта колонок листа Итого",
        "",
        "| column_index | column_name | block_name | formula_example | source_sheet | source_range_or_reference | source_type | target_db_table | target_db_field | automation_status | notes |",
        "| ---: | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- |",
    ]
    for column in columns:
        lines.append(
            "| {column_index} | {column_name} | {block_name} | `{formula}` | {source_sheet} | `{refs}` | {source_type} | {target_db_table} | {target_db_field} | {automation_status} | {notes} |".format(
                column_index=column["column_index"],
                column_name=markdown_escape(column["column_name"]),
                block_name=markdown_escape(column["block_name"]),
                formula=markdown_escape(column["formula_example"]),
                source_sheet=markdown_escape(column["source_sheet"] or "-"),
                refs=markdown_escape(column["source_range_or_reference"]),
                source_type=column["source_type"],
                target_db_table=column["target_db_table"],
                target_db_field=column["target_db_field"],
                automation_status=column["automation_status"],
                notes=markdown_escape(column["notes"]),
            )
        )
    return lines


def build_formula_map(columns: list[dict[str, object]], summary: dict[str, object]) -> str:
    by_block: dict[str, list[dict[str, object]]] = defaultdict(list)
    for column in columns:
        if column["formula_example"]:
            by_block[column["block_name"]].append(column)

    lines = [
        "# TOTAL_REPORT_FORMULA_MAP",
        "",
        "## Что такое лист Итого",
        "",
        "- `Итого` — это не raw-выгрузка, а широкая сводная витрина Excel с примесью ручных полей, wide-pivot блоков и локальных формул.",
        "- Большая часть структуры листа завязана на другие вкладки workbook, а не на единый нормализованный источник.",
        "- Для database-first архитектуры лист надо разложить на: `fact_*` таблицы источников, `dim_product`, и итоговую витрину `mart_total_report`.",
        "",
        "## Главные формульные паттерны",
        "",
    ]

    ordered_blocks = [spec.name for spec in BLOCK_SPECS] + ["Прочее / требует ручной группировки"]
    for block_name in ordered_blocks:
        sample_columns = by_block.get(block_name)
        if not sample_columns:
            continue
        lines.append(f"### {block_name}")
        lines.append("")
        seen_examples: set[str] = set()
        added = 0
        for column in sample_columns:
            formula = column["formula_example"]
            if formula in seen_examples:
                continue
            seen_examples.add(formula)
            lines.append(
                f"- `{column['column_name']}`: `{formula}`. {column['formula_pattern']}"
            )
            added += 1
            if added >= 5:
                break
        lines.append("")

    lines.extend(
        [
            "## Зависимости по листам",
            "",
            "- `Воронка на день`: базовый источник по дням и nm_id для показов, переходов, корзин, заказов, CTR и части локальных KPI.",
            "- `РасходРК`: источник рекламных расходов и wide-матрицы рекламных кампаний; часть полей Итого использует `SUMIF/SUMIFS` по этому листу.",
            "- `РК стата`: в текущем xlsm служит как целевой рекламный блок, но в итоговой архитектуре должен читаться как `fact_ad_campaign_nm_day`; для production нужен лаг минимум D-2.",
            "- `ВБро`: ручной внешний сервис; в Итого сейчас стоят прямые ссылки на лист `ВБро`, но автоматического расчета прибыли делать не нужно.",
            "- `Точка вх`: широкий pivot по surfaces/customer profile; нужен CSV/Excel export или private endpoint.",
            "- `Локализация`: wide-региональные показатели; текущий `region-sale` покрывает только часть задачи, целевой источник — `orders-geography`.",
            "- `Сравнение карточек`: внешний comparison/MPStat блок, часть формул Итого читает его как справочник рейтингов и позиций.",
            "- `Поисковые запросы`: wide-матрица поисковых фраз; в БД должна стать `fact_search_query_metric`, а конкурентные percentile поля остаются отдельной зоной подтверждения.",
            "",
            "## Что уже можно перенести в БД",
            "",
            "- Уже покрывается текущим API-пайплайном: базовая воронка, расходы рекламы, часть рекламной статистики, часть поисковых запросов, часть локализации.",
            "- Требует manual/export: `ВБро`, `Точка вх`, часть `Локализация`, часть wide campaign/search blocks, которые в xlsm пока построены как Excel-pivot.",
            "- Требует external service: `Сравнение карточек` / MPStat-подобные поля.",
            "- Требует отдельного подтверждения: wide-формулы с `#REF!`, скрытая макросная логика и бизнес-смысл некоторых промежуточных KPI колонок.",
            "",
            "## Риски при переносе",
            "",
            f"- Колонки с `#REF!`: `{summary['ref_error_columns']}`. Эти формулы нельзя переносить без ручной проверки.",
            f"- Колонки с фиксированными диапазонами: `{summary['fixed_range_columns']}`. Они сломаются при масштабировании списка артикулов/кампаний.",
            "- В workbook есть `vbaProject.bin`: часть подготовки данных может выполняться макросами, что в этом анализе не исполнялось.",
        ]
    )
    return "\n".join(lines)


def write_csv(columns: list[dict[str, object]]) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with CSV_PATH.open("w", encoding="utf-8-sig", newline="") as file_handle:
        writer = csv.DictWriter(
            file_handle,
            fieldnames=[
                "column_index",
                "column_name",
                "block_name",
                "formula",
                "source_sheet",
                "source_type",
                "target_db_table",
                "automation_status",
                "needs_confirmation",
            ],
        )
        writer.writeheader()
        for column in columns:
            writer.writerow(
                {
                    "column_index": column["column_index"],
                    "column_name": column["column_name"],
                    "block_name": column["block_name"],
                    "formula": column["formula_example"],
                    "source_sheet": column["source_sheet"],
                    "source_type": column["source_type"],
                    "target_db_table": column["target_db_table"],
                    "automation_status": column["automation_status"],
                    "needs_confirmation": column["needs_confirmation"],
                }
            )


def build_spec_text(
    workbook_path: Path,
    sheet_infos: list[SheetInfo],
    summary: dict[str, object],
    zip_info: dict[str, object],
    columns: list[dict[str, object]],
) -> str:
    lines = [
        "# TOTAL_REPORT_SPEC",
        "",
        f"- Workbook: `{workbook_path.name}`",
        f"- Workbook path: `{workbook_path}`",
        f"- Total sheets: `{len(sheet_infos)}`",
        f"- VBA project detected: `{zip_info['has_vba']}`",
        f"- External link parts in xlsm zip: `{len(zip_info['external_links'])}`",
        f"- Sheet `Итого`: `{summary['max_row']}` rows x `{summary['used_column_count']}` used columns",
        f"- Last non-empty column on `Итого`: `{summary['last_nonempty_column_index']} ({get_column_letter(summary['last_nonempty_column_index'])})`",
        f"- Formula columns on `Итого`: `{summary['formula_columns']}`",
        f"- Cross-sheet formula columns on `Итого`: `{summary['cross_sheet_columns']}`",
        f"- Columns with fixed ranges: `{summary['fixed_range_columns']}`",
        f"- Columns with `#REF!` in sample formula: `{summary['ref_error_columns']}`",
        "",
    ]
    lines.extend(build_workbook_structure_table(sheet_infos))
    lines.append("")
    lines.extend(build_source_dependencies_section(summary))
    lines.append("")
    lines.extend(build_block_summary_section(summary))
    lines.append("")
    lines.extend(build_risk_section(summary, zip_info))
    lines.append("")
    lines.extend(build_column_table(columns))
    return "\n".join(lines)


def main() -> None:
    workbook_path = resolve_workbook_path()
    workbook, sheet_infos, zip_info = load_workbook_info(workbook_path)
    columns, summary = analyze_total_sheet(workbook)

    DOCS_DIR.mkdir(parents=True, exist_ok=True)
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    SPEC_PATH.write_text(
        build_spec_text(workbook_path, sheet_infos, summary, zip_info, columns),
        encoding="utf-8",
    )
    FORMULA_MAP_PATH.write_text(build_formula_map(columns, summary), encoding="utf-8")
    write_csv(columns)

    result = {
        "workbook": workbook_path.name,
        "sheet_count": len(sheet_infos),
        "itogo_columns": summary["used_column_count"],
        "itogo_formula_columns": summary["formula_columns"],
        "itogo_cross_sheet_columns": summary["cross_sheet_columns"],
        "source_sheets": summary["source_sheets"],
        "status_usage": summary["status_usage"],
        "vba_detected": zip_info["has_vba"],
        "external_links": zip_info["external_links"],
        "outputs": {
            "spec": str(SPEC_PATH),
            "formula_map": str(FORMULA_MAP_PATH),
            "csv": str(CSV_PATH),
        },
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
