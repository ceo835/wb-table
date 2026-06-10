from __future__ import annotations

from decimal import Decimal

from openpyxl import Workbook

from src.importers.entry_points_importer import (
    ENTRY_POINT_CONFLICT_COLUMNS,
    build_entry_point_summary,
    parse_entry_points_xlsx,
    prepare_entry_point_upsert_rows,
)


def _build_entry_points_workbook(path) -> None:
    workbook = Workbook()
    info_sheet = workbook.active
    info_sheet.title = "Общая информация"
    info_sheet["A1"] = "Дата отчёта"
    info_sheet["B1"] = "2026-06-07"

    workbook.create_sheet("Детализация по точкам входа")
    data_sheet = workbook.create_sheet("Детализация по артикулам")
    data_sheet.append(["Детальный отчет по точкам входа"])
    data_sheet.append(
        [
            "Раздел",
            "Точка входа",
            "Артикул ВБ",
            "Артикул продавца",
            "Бренд",
            "Название",
            "Предмет",
            "Показы",
            "Переходы в карточку",
            "CTR",
            "Добавления в корзину",
            "Конверсия в корзину",
            "Заказы",
            "Конверсия в заказ",
        ]
    )
    data_sheet.append(
        [
            "Поиск",
            "Выдача",
            197330807,
            "BlackWOM5",
            "PALEY",
            "Трусы",
            "Трусы",
            1000,
            50,
            0.05,
            10,
            0.2,
            4,
            0.4,
        ]
    )
    data_sheet["J3"].number_format = "0.00%"
    data_sheet["L3"].number_format = "0.00%"
    data_sheet["N3"].number_format = "0.00%"
    workbook.save(path)


def test_parse_entry_points_xlsx_detects_main_sheet_and_normalizes_metrics(tmp_path):
    file_path = tmp_path / "entry_points_export_2026-06-07.xlsx"
    _build_entry_points_workbook(file_path)

    parsed = parse_entry_points_xlsx(str(file_path))

    assert parsed.detected_date.isoformat() == "2026-06-07"
    assert parsed.sheet_name == "Детализация по артикулам"
    assert parsed.rows_read == 1
    assert not parsed.missing_required_columns
    row = parsed.rows_normalized[0]
    assert row["nm_id"] == 197330807
    assert row["section"] == "Поиск"
    assert row["entry_point"] == "Выдача"
    assert row["impressions"] == Decimal("1000")
    assert row["card_clicks"] == Decimal("50")
    assert row["ctr"] == Decimal("5.00")
    assert row["add_to_cart_conversion"] == Decimal("20.0")
    assert row["order_conversion"] == Decimal("40.0")
    assert row["source_status"] == "CSV_EXPORT"


def test_parse_entry_points_xlsx_releases_file_handle(tmp_path):
    file_path = tmp_path / "entry_points_export_2026-06-07.xlsx"
    _build_entry_points_workbook(file_path)

    parse_entry_points_xlsx(str(file_path))
    file_path.unlink()

    assert not file_path.exists()


def test_prepare_entry_point_upsert_rows_deduplicates_by_date_nm_section_entry_point():
    rows = prepare_entry_point_upsert_rows(
        [
            {
                "date": "2026-06-07",
                "nm_id": 197330807,
                "section": "Поиск",
                "entry_point": "Выдача",
                "order_count": Decimal("1"),
            },
            {
                "date": "2026-06-07",
                "nm_id": 197330807,
                "section": "Поиск",
                "entry_point": "Выдача",
                "order_count": Decimal("5"),
            },
        ]
    )
    assert ENTRY_POINT_CONFLICT_COLUMNS == ("date", "nm_id", "section", "entry_point")
    assert len(rows) == 1
    assert rows[0]["order_count"] == Decimal("5")
    assert rows[0]["loaded_at"] is not None


def test_build_entry_point_summary_includes_target_table_and_source_status(tmp_path):
    file_path = tmp_path / "entry_points_export_2026-06-07.xlsx"
    _build_entry_points_workbook(file_path)

    parsed = parse_entry_points_xlsx(str(file_path))
    summary = build_entry_point_summary(parsed)

    assert summary["target_table"] == "fact_entry_point_day"
    assert summary["source_status"] == "CSV_EXPORT"
